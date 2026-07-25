"""
Hapu Mail — Đăng nhập YouTube/Google vào GPM-Login profile qua proxy KiotProxy
==========================================================================================
Mô hình: 1 KEY kiotproxy = 1 luồng (thread).
  - Mỗi luồng xử lý tuần tự các tài khoản thuộc key đó.
  - Trước mỗi tài khoản: gọi API kiotproxy lấy/đổi IP mới (chờ tới lượt đổi nếu cần),
    ghi proxy vào GPM profile, test proxy, mở profile rồi tự đăng nhập Google
    (tái sử dụng logic đăng nhập của auto_login_tool.py).
  - Nhiều key → nhiều luồng chạy song song.

Yêu cầu: GPM-Login đang chạy, đặt cùng thư mục với auto_login_tool.py và gpm_client.py.
Chạy:  python kiot_login_tool.py
"""
import asyncio
import sys

# ── ÉP stdout/stderr DÙNG UTF-8 để tiếng Việt có dấu KHÔNG bị thành '?' trên console
#    Windows. reconfigure() nhiều khi thất bại âm thầm trên Python Store/Windows Terminal,
#    nên ta ép MẠNH: (1) đặt code page console = 65001 (UTF-8) qua ctypes, (2) BỌC LẠI
#    stdout/stderr bằng TextIOWrapper trên .buffer với encoding utf-8.
try:
    import ctypes as _ctypes
    _ctypes.windll.kernel32.SetConsoleOutputCP(65001)
    _ctypes.windll.kernel32.SetConsoleCP(65001)
except Exception:
    pass
try:
    import io as _io
    if getattr(sys.stdout, "buffer", None) is not None:
        sys.stdout = _io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8",
                                       errors="replace", line_buffering=True)
    if getattr(sys.stderr, "buffer", None) is not None:
        sys.stderr = _io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8",
                                       errors="replace", line_buffering=True)
except Exception:
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

# ── GHI LOG RA FILE run_log.txt (UTF-8) ngay từ Python, không lệ thuộc redirect của .bat.
#    Nhờ vậy chạy WINDOWLESS (pythonw, không có console) vẫn có log để soi lỗi sau này.
import os as _os_early
class _Tee:
    def __init__(self, *streams):
        self.streams = [s for s in streams if s is not None]
    def write(self, s):
        for st in self.streams:
            try:
                st.write(s); st.flush()
            except Exception:
                pass
    def flush(self):
        for st in self.streams:
            try:
                st.flush()
            except Exception:
                pass
    def isatty(self):
        return False
try:
    _logpath = _os_early.path.join(_os_early.path.dirname(_os_early.path.abspath(__file__)),
                                   "run_log.txt")
    _logf = open(_logpath, "w", encoding="utf-8", errors="replace", buffering=1)
    sys.stdout = _Tee(sys.stdout, _logf)
    sys.stderr = _Tee(sys.stderr, _logf)
except Exception:
    pass

import threading
import time
import traceback
import queue as _queue          # hàng đợi account/task dùng chung cho mọi luồng
import tkinter as tk
from collections import OrderedDict
from tkinter import font as tkfont
from tkinter import filedialog, messagebox, scrolledtext, ttk

import os
from pathlib import Path

# stdout/stderr UTF-8 để print tiếng Việt (vd vị trí proxy) không làm crash log
for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

import requests

# Tái sử dụng logic đăng nhập + helper GPM đã có sẵn
from auto_login_tool import (do_google_login, do_create_2fa, do_change_2fa,
                             do_channel, do_change_banner, do_leave_admin,
                             do_add_brand_admin, do_change_avatar, do_taodata_images,
                             do_accept_brand_invite,
                             gpm_list, gpm_start, gpm_stop,
                             sync_time_offset, _DeviceCodeChallenge)
from gpm_client import GPMClient
try:
    import gsheet_client
except Exception:
    gsheet_client = None

# ══════════════════════════════════════════════════════════════════════════
#  PHIÊN BẢN TOOL — MỖI LẦN SỬA CODE, ĐỔI SỐ NÀY (ngày + số thứ tự trong ngày).
#  Hiện ở tiêu đề cửa sổ + header để mỗi máy biết đang chạy bản nào.
# ══════════════════════════════════════════════════════════════════════════
APP_VERSION = "2026.07.25-a"

# ══════════════════════════════════════════════════════════════════════════
#  TỰ ĐỘNG CẬP NHẬT (qua GitHub) — mỗi máy khi mở tool sẽ hỏi version.json trên
#  GitHub; nếu khác APP_VERSION → hiện nút "Cập nhật ngay" ở header (KHÔNG tự tải).
#  ⚠ SAU KHI TẠO REPO GITHUB: đổi USER/REPO bên dưới cho đúng (giữ nhánh 'main').
#     Cấu trúc repo cần có:  version.json  +  CHUONG_TRINH/auto_login_tool.py
#                            +  CHUONG_TRINH/kiot_login_tool.py
#     version.json ví dụ: {"version":"2026.07.24-a","note":"...",
#                          "files":["CHUONG_TRINH/auto_login_tool.py",
#                                   "CHUONG_TRINH/kiot_login_tool.py"]}
# ══════════════════════════════════════════════════════════════════════════
UPDATE_BASE_URL = "https://raw.githubusercontent.com/TuanHapuMedia/Hapu-ToolLogin/main"

# Nạp config.env (nếu có) để lấy sẵn KIOT_API_TOKEN
try:
    from dotenv import load_dotenv
    # override=True: token đã lưu trong config.env của máy này LUÔN thắng (không bị biến
    # môi trường hệ thống ghi đè) → đúng token của từng máy.
    load_dotenv(Path(__file__).parent / "config.env", override=True)
except Exception:
    pass

# ──────────────────────────────────────────────────────────
KIOT_ROOT = "https://api.kiotproxy.com"
REGIONS = ["random", "bac", "trung", "nam"]
# Lock dùng chung khi ghi SQLite proxy của GPM (tránh nhiều luồng ghi cùng lúc)
DB_LOCK = threading.Lock()
# Lock khi tạo profile GPM bằng tự động click (chỉ 1 luồng tạo tại 1 thời điểm)
GPM_UI_LOCK = threading.Lock()
# ──────────────────────────────────────────────────────────


class KiotError(Exception):
    """Lỗi trả về từ API KiotProxy."""


class KiotProxyClient:
    """Client gọi API KiotProxy (https://api.kiotproxy.com)."""

    def __init__(self, root: str = KIOT_ROOT, timeout: int = 30):
        self.root = root.rstrip("/")
        self.timeout = timeout
        self.session = requests.Session()

    # ── low level ──────────────────────────────────────────
    def _get(self, path: str, params: dict = None, headers: dict = None) -> dict:
        url = f"{self.root}/{path.lstrip('/')}"
        r = self.session.get(url, params=params or {}, headers=headers or {}, timeout=self.timeout)
        try:
            d = r.json()
        except Exception:
            raise KiotError(f"Phản hồi không phải JSON (HTTP {r.status_code}): {r.text[:120]!r}")
        if not d.get("success", False):
            msg = d.get("message") or d.get("error") or f"code={d.get('code')}"
            raise KiotError(msg)
        return d

    @staticmethod
    def _parse(data: dict) -> dict:
        """Chuẩn hoá 1 bản ghi proxy trả về."""
        return {
            "host": data.get("host", ""),
            "http_port": data.get("httpPort"),
            "socks5_port": data.get("socks5Port"),
            "http": data.get("http", ""),
            "socks5": data.get("socks5", ""),
            "real_ip": data.get("realIpAddress", ""),
            "location": data.get("location", ""),
            "ttl": data.get("ttl"),
            "ttc": data.get("ttc"),
            # nextRequestAt: mốc thời gian (ms) được phép đổi proxy tiếp theo
            "next_at": data.get("nextRequestAt"),
        }

    # ── V1 proxy endpoints (xác thực bằng proxy key) ───────
    def get_new(self, key: str, region: str = "random") -> dict:
        """GET /api/v1/proxies/new — lấy proxy mới / đổi IP (khi key đã tới lượt đổi)."""
        d = self._get("api/v1/proxies/new", {"key": key, "region": region})
        return self._parse(d.get("data", {}))

    def get_current(self, key: str) -> dict:
        """GET /api/v1/proxies/current — thông tin proxy hiện tại của key."""
        d = self._get("api/v1/proxies/current", {"key": key})
        return self._parse(d.get("data", {}))

    def out(self, key: str) -> bool:
        """GET /api/v1/proxies/out — thoát proxy khỏi key."""
        d = self._get("api/v1/proxies/out", {"key": key})
        return bool(d.get("data", False))

    # ── Public endpoints (xác thực bằng API Token qua header) ──
    def list_keys(self, api_token: str) -> list:
        """
        GET /api/public/keys — liệt kê tất cả key trong tài khoản.
        Trả về list dict: {value, expiration, status, description, id}
        """
        d = self._get("api/public/keys", headers={"Api-token": api_token})
        out = []
        for k in d.get("data", []) or []:
            out.append({
                "id": k.get("id", ""),
                "value": k.get("value", ""),
                "status": k.get("status", ""),
                "description": k.get("description", "") or "",
                "expiration": k.get("expirationAt", ""),
            })
        return out

    def user_info(self, api_token: str) -> dict:
        """GET /api/public/get-user-info — thông tin tài khoản (username, balance...)."""
        d = self._get("api/public/get-user-info", headers={"Api-token": api_token})
        return d.get("data", {}) or {}


# ═══════════════════════════════════════════════════════════
# GUI
# ═══════════════════════════════════════════════════════════
class KiotLoginApp:
    COLOR_BG = "#1e1e2e"
    COLOR_PANEL = "#2a2a3e"
    COLOR_ACCENT = "#0ea5e9"
    COLOR_GREEN = "#22c55e"
    COLOR_RED = "#ef4444"
    COLOR_YELLOW = "#f59e0b"
    COLOR_TEXT = "#e2e8f0"
    COLOR_MUTED = "#94a3b8"

    def __init__(self):
        self.root = tk.Tk()
        self.root.title(f"Hapu Mail — bản {APP_VERSION}")
        self.root.geometry("960x720")
        self.root.configure(bg=self.COLOR_BG)
        self.root.resizable(True, True)

        self.stop_flag = False
        self.profiles: list[dict] = []
        self.kiot = KiotProxyClient()
        self._threads: list[threading.Thread] = []
        self.results: list[dict] = []          # kết quả từng account
        self._results_lock = threading.Lock()
        self._progress_lock = threading.RLock()  # khóa đồng bộ tiến độ (đa luồng)
        self.available_keys: list[str] = []    # key tải từ API token
        self._acc_creds: dict = {}             # lưu creds để xuất CSV đầy đủ
        self._active_workers: int = 0          # số worker đang chạy (dynamic)
        self._task_queue = None                # queue task hiện tại
        self._acc_queue = None                 # hàng đợi account chung (mọi luồng cùng rút)
        self._progress_total: int = 0          # tổng account cho lần chạy
        self._progress_done: int = 0           # đã xử lý xong
        self._stt_counter: int = 0             # STT bảng kết quả
        self.task_vars: dict = {}              # BooleanVar cho từng tác vụ
        self._task_row_frames: dict = {}       # frame row tác vụ (để highlight)
        self._task_order: list = []            # THỨ TỰ người dùng click chọn tác vụ

        self._build_ui()
        self._load_profiles()
        # Nếu đã có API token (từ config.env) -> tự nạp key luôn khi mở tool
        if self.token_var.get().strip():
            self.root.after(600, self._load_keys_from_token)
        # CHẾ ĐỘ TEST: CHỈ khi có phiên test đang DỞ (tồn tại file tiến độ) mới nạp lại
        # list account + bật lại chế độ test. Chạy bình thường không bị ảnh hưởng.
        try:
            if self._progress_path().exists() and self._session_path().exists():
                if self._session_load_accounts():
                    self.test_mode_var.set(True)
                    self.root.after(50, self._update_line_count)
        except Exception:
            pass

        # KIỂM TRA BẢN MỚI (nền, không chặn UI). Chỉ BÁO + hiện nút, KHÔNG tự tải.
        self._update_info = None
        self.root.after(1500, self._check_update_async)

    # ══════════ TỰ ĐỘNG CẬP NHẬT (GitHub) ══════════
    def _check_update_async(self):
        """Hỏi version.json trên GitHub ở luồng nền; nếu khác APP_VERSION → hiện nút cập nhật."""
        if "USER/REPO" in UPDATE_BASE_URL:
            return   # chưa cấu hình repo → bỏ qua, không báo lỗi
        def _work():
            try:
                r = requests.get(f"{UPDATE_BASE_URL}/version.json", timeout=8)
                if r.status_code != 200:
                    return
                info = r.json()
                latest = str(info.get("version", "")).strip()
                if latest and latest != APP_VERSION:
                    self._update_info = info
                    self.root.after(0, self._show_update_banner)
            except Exception:
                pass   # mạng lỗi / không có repo → im lặng
        threading.Thread(target=_work, daemon=True).start()

    def _show_update_banner(self):
        try:
            latest = str(self._update_info.get("version", "")).strip()
            self._update_btn.config(text=f"⬇ Có bản mới {latest} — Cập nhật ngay")
            self._update_btn.pack(side=tk.RIGHT, padx=6, pady=8)
        except Exception:
            pass

    def _do_update(self):
        """BẠN bấm nút → tải file mới từ GitHub, kiểm tra hợp lệ, thay file, báo khởi động lại."""
        info = self._update_info or {}
        latest = str(info.get("version", "")).strip()
        # Trên GitHub để file PHẲNG ở gốc (vd 'auto_login_tool.py'); tool tự ghi vào CHUONG_TRINH/.
        files = info.get("files") or ["auto_login_tool.py", "kiot_login_tool.py"]
        note = str(info.get("note", "")).strip()
        if not messagebox.askyesno(
                "Cập nhật tool",
                f"Tải bản {latest} và thay file?\n\n{note}\n\n"
                "• Cấu hình riêng (config.env, data, 2fa_recovery) KHÔNG bị đụng.\n"
                "• Xong cần TẮT rồi mở lại tool."):
            return
        self._update_btn.config(state=tk.DISABLED, text="⏳ Đang tải bản mới…")

        def _work():
            try:
                root = Path(__file__).parent.parent   # thư mục gốc tool (chứa CHUONG_TRINH)
                staged = {}
                for rel in files:
                    rel = str(rel).strip().strip("/")
                    # CHẶN đường dẫn nguy hiểm
                    if ".." in rel or rel.startswith("/") or ":" in rel:
                        raise ValueError(f"Đường dẫn file không hợp lệ: {rel}")
                    rr = requests.get(f"{UPDATE_BASE_URL}/{rel}", timeout=30)
                    rr.raise_for_status()
                    data = rr.content
                    if not data or len(data) < 50:
                        raise ValueError(f"File tải về rỗng/hỏng: {rel}")
                    # File .py: BẮT BUỘC compile OK trước khi thay (tránh hỏng tool)
                    if rel.endswith(".py"):
                        compile(data.decode("utf-8"), rel, "exec")
                    staged[rel] = data
                # Sao lưu + ghi đè. File .py PHẲNG trên GitHub → ghi vào CHUONG_TRINH/ trên máy;
                # nếu rel đã có sẵn thư mục (chứa '/') thì giữ nguyên theo rel.
                for rel, data in staged.items():
                    if "/" in rel:
                        p = root / rel
                    elif rel.endswith(".py"):
                        p = root / "CHUONG_TRINH" / rel
                    else:
                        p = root / rel
                    p.parent.mkdir(parents=True, exist_ok=True)
                    if p.exists():
                        try:
                            (p.parent / (p.name + ".bak")).write_bytes(p.read_bytes())
                        except Exception:
                            pass
                    p.write_bytes(data)
                self.root.after(0, lambda: [
                    messagebox.showinfo(
                        "Cập nhật xong",
                        f"Đã cập nhật bản {latest} ({len(staged)} file).\n\n"
                        "Hãy TẮT HẲN tool rồi mở lại bằng CHAY_TOOL.bat."),
                    self._update_btn.config(state=tk.NORMAL, text="✓ Đã cập nhật — khởi động lại")])
            except Exception as e:
                _msg = str(e)[:200]
                self.root.after(0, lambda: [
                    messagebox.showerror("Lỗi cập nhật",
                                         f"Không cập nhật được (đã giữ nguyên bản cũ):\n{_msg}"),
                    self._update_btn.config(state=tk.NORMAL,
                                            text=f"⬇ Có bản mới {latest} — Thử lại")])
        threading.Thread(target=_work, daemon=True).start()

    # ── UI ─────────────────────────────────────────────────
    def _build_ui(self):
        bold14 = tkfont.Font(family="Segoe UI", size=14, weight="bold")
        norm10 = tkfont.Font(family="Segoe UI", size=10)
        small9 = tkfont.Font(family="Segoe UI", size=9)
        mono10 = tkfont.Font(family="Consolas", size=10)
        mono9 = tkfont.Font(family="Consolas", size=9)

        # ── Header ──
        hdr = tk.Frame(self.root, bg=self.COLOR_ACCENT, height=48)
        hdr.pack(fill=tk.X)
        tk.Label(
            hdr, text="  🌐  HAPU MAIL  —  Quản lý tài khoản Google",
            font=bold14, bg=self.COLOR_ACCENT, fg="white", pady=10
        ).pack(side=tk.LEFT)
        # Nhãn phiên bản (góc phải header) — biết ngay máy đang chạy bản nào
        tk.Label(
            hdr, text=f"bản {APP_VERSION}   ",
            font=tkfont.Font(family="Segoe UI", size=9),
            bg=self.COLOR_ACCENT, fg="#dbeafe", pady=14
        ).pack(side=tk.RIGHT)
        # Nút CẬP NHẬT (ẩn; chỉ hiện khi phát hiện bản mới trên GitHub)
        self._update_btn = tk.Button(
            hdr, text="", font=tkfont.Font(family="Segoe UI", size=9, weight="bold"),
            bg="#f59e0b", fg="white", bd=0, cursor="hand2", padx=10,
            activebackground="#d97706", activeforeground="white",
            command=self._do_update)
        # chưa pack — _show_update_banner() sẽ pack khi có bản mới

        # ── Notebook chính 2 tab ──
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Main.TNotebook",
                        background=self.COLOR_BG,
                        borderwidth=0, tabmargins=0)
        style.configure("Main.TNotebook.Tab",
                        background="#16162a", foreground="#9090c0",
                        font=("Segoe UI", 10), padding=[18, 8])
        style.map("Main.TNotebook.Tab",
                  background=[("selected", self.COLOR_ACCENT)],
                  foreground=[("selected", "white")])

        self.main_nb = ttk.Notebook(self.root, style="Main.TNotebook")
        self.main_nb.pack(fill=tk.BOTH, expand=True)

        # ── Tab 1: Thiết lập ──────────────────────────────────
        tab1 = tk.Frame(self.main_nb, bg=self.COLOR_BG)
        self.main_nb.add(tab1, text="⚙  Thiết lập")

        body = tk.Frame(tab1, bg=self.COLOR_BG)
        body.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)

        # ── Panel tác vụ (cột trái) ──
        TASK_BG      = "#13132a"
        TASK_SEL     = "#1e2050"
        TASK_BORDER  = "#2e2e55"
        bold10 = tkfont.Font(family="Segoe UI", size=10, weight="bold")

        task_panel = tk.Frame(body, bg=TASK_BG, width=170)
        task_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        task_panel.pack_propagate(False)

        tk.Label(task_panel, text="TÁC VỤ", font=bold10,
                 bg=TASK_BG, fg="#6b7aad", pady=6, padx=10,
                 anchor=tk.W).pack(fill=tk.X)
        tk.Frame(task_panel, bg=TASK_BORDER, height=1).pack(fill=tk.X)

        # Danh sách tác vụ: (id, emoji, tên, đã hỗ trợ)
        _TASK_DEFS = [
            ("login",      "🔑", "Đăng nhập",      True),
            ("change2fa",  "🔐", "Đổi 2FA",         False),
            ("create2fa",  "✨", "Tạo 2FA",          False),
            ("channel",    "📺", "Tạo kênh",        False),
            ("getlink",    "🔗", "Lấy link kênh",   False),
            ("banner",     "🎨", "Thay ảnh bìa",    True),
            ("rmadmin",    "🚪", "Thoát quản trị",  True),
            ("addqtth",    "➕", "Add QT TH",       False),
            ("addqtonly",  "➕", "Add Thêm QT",     False),
            ("cnqt",       "✅", "Chấp nhận QT",   False),
            ("taodata",    "🗂️", "Tạo Data",         False),
        ]

        def _make_task_row(parent, tid, icon, label, ready):
            var = tk.BooleanVar(value=(tid == "login"))
            self.task_vars[tid] = var

            row = tk.Frame(parent, bg=TASK_SEL if tid == "login" else TASK_BG,
                           cursor="hand2")
            row.pack(fill=tk.X, pady=1)
            self._task_row_frames[tid] = row

            inner = tk.Frame(row, bg=row["bg"])
            inner.pack(fill=tk.X, padx=8, pady=5)

            # Checkbox hình vuông tự vẽ
            chk_canvas = tk.Canvas(inner, width=16, height=16, bg=inner["bg"],
                                   highlightthickness=0, cursor="hand2")
            chk_canvas.pack(side=tk.LEFT, padx=(0, 6))

            # Icon để RIÊNG 1 Label, dùng font "Segoe UI Emoji" (hiển thị emoji màu chuẩn) +
            # cố định bề rộng → mọi hàng thẳng đều, không bị xô.
            icon_lbl = tk.Label(inner, text=icon, width=2,
                                font=tkfont.Font(family="Segoe UI Emoji", size=11),
                                bg=inner["bg"], fg="#c9c9ff", anchor=tk.CENTER, cursor="hand2")
            icon_lbl.pack(side=tk.LEFT, padx=(0, 4))

            lbl = tk.Label(inner, text=label,
                           font=tkfont.Font(family="Segoe UI", size=10,
                                            weight="bold" if tid == "login" else "normal"),
                           bg=inner["bg"],
                           fg="white" if tid == "login" else "#9090c0",
                           anchor=tk.W, cursor="hand2")
            lbl.pack(side=tk.LEFT, fill=tk.X, expand=True)

            def _draw_check(canvas, checked, bg):
                canvas.delete("all")
                canvas.config(bg=bg)
                r = 3
                canvas.create_rectangle(0, 0, 15, 15,
                    outline=self.COLOR_ACCENT if checked else "#4a4a70",
                    fill=self.COLOR_ACCENT if checked else "#0d0d1a",
                    width=2)
                if checked:
                    canvas.create_line(3, 8, 6, 12, 12, 4,
                                       fill="white", width=2, capstyle="round",
                                       joinstyle="round")

            def _apply_row_state(t, is_on):
                """Cập nhật giao diện 1 row theo trạng thái is_on."""
                v   = self.task_vars[t]
                rw  = self._task_row_frames[t]
                # tìm label và canvas trong row
                for inner_w in rw.winfo_children():
                    inner_w.config(bg=TASK_SEL if is_on else TASK_BG)
                    _lbl = _cv = None
                    for child in inner_w.winfo_children():
                        try:
                            child.config(bg=TASK_SEL if is_on else TASK_BG)
                        except Exception:
                            pass
                        if isinstance(child, tk.Label):
                            _lbl = child
                        elif isinstance(child, tk.Canvas):
                            _cv = child
                    if _lbl:
                        _lbl.config(
                            fg="white" if is_on else "#9090c0",
                            font=tkfont.Font(family="Segoe UI", size=10,
                                             weight="bold" if is_on else "normal")
                        )
                    if _cv:
                        _draw_check(_cv, is_on, TASK_SEL if is_on else TASK_BG)
                rw.config(bg=TASK_SEL if is_on else TASK_BG)
            # LƯU tham chiếu để nơi khác (khôi phục phiên test) vẽ lại checkbox cho khớp biến.
            self._apply_row_state_fn = _apply_row_state

            def _toggle(t=tid, v=var, rw=row, lbl=lbl, cv=chk_canvas):
                new_state = not v.get()
                v.set(new_state)
                # GHI THỨ TỰ CLICK: bật → thêm vào cuối; tắt → bỏ khỏi danh sách.
                if new_state:
                    if t not in self._task_order:
                        self._task_order.append(t)
                else:
                    if t in self._task_order:
                        self._task_order.remove(t)
                is_on = new_state
                bg = TASK_SEL if is_on else TASK_BG
                rw.config(bg=bg)
                for child in rw.winfo_children():
                    child.config(bg=bg)
                    for gc in child.winfo_children():
                        try:
                            gc.config(bg=bg)
                        except Exception:
                            pass
                lbl.config(
                    fg="white" if is_on else "#9090c0",
                    font=tkfont.Font(family="Segoe UI", size=10,
                                     weight="bold" if is_on else "normal")
                )
                _draw_check(cv, is_on, bg)
                self._refresh_start_btn()

            _draw_check(chk_canvas, var.get(), TASK_SEL if tid == "login" else TASK_BG)

            for widget in (row, inner, chk_canvas, icon_lbl, lbl):
                widget.bind("<Button-1>", lambda e, fn=_toggle: fn())

        for tid, icon, label, ready in _TASK_DEFS:
            _make_task_row(task_panel, tid, icon, label, ready)

        tk.Frame(task_panel, bg=TASK_BORDER, height=1).pack(fill=tk.X, pady=(4, 0))
        tk.Label(task_panel,
                 text="Chạy theo ĐÚNG THỨ TỰ\nbạn CLICK chọn",
                 font=tkfont.Font(family="Segoe UI", size=8),
                 bg=TASK_BG, fg="#5050a0",
                 justify=tk.LEFT, padx=10).pack(anchor=tk.W, pady=4)

        # ── Cột giữa: Account input ──
        left = tk.Frame(body, bg=self.COLOR_BG)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 6))

        tk.Label(left, text="Tài khoản",
                 font=norm10, bg=self.COLOR_BG, fg=self.COLOR_MUTED).pack(anchor=tk.W)
        tk.Label(
            left,
            text=("Mỗi dòng:  mail | matkhau | recovery | 2fa_secret\n"
                  "Bấm '⬇ Tải keys' để tool TỰ CHIA key cho account "
                  "(không cần nhập key)."),
            font=mono9, justify=tk.LEFT,
            bg=self.COLOR_BG, fg=self.COLOR_ACCENT).pack(anchor=tk.W, pady=(0, 4))

        # Hàng nút nhập từ file / xoá
        imp_row = tk.Frame(left, bg=self.COLOR_BG)
        imp_row.pack(fill=tk.X, pady=(0, 4))
        tk.Button(
            imp_row, text="📂 Nhập từ file (xlsx / csv / txt)",
            font=small9, bg=self.COLOR_PANEL, fg=self.COLOR_TEXT,
            relief=tk.FLAT, cursor="hand2", activebackground=self.COLOR_ACCENT,
            command=self._import_from_file, pady=3
        ).pack(side=tk.LEFT)
        tk.Button(
            imp_row, text="🗑 Xoá ô nhập",
            font=small9, bg=self.COLOR_PANEL, fg=self.COLOR_TEXT,
            relief=tk.FLAT, cursor="hand2", activebackground=self.COLOR_RED,
            command=self._show_acc_placeholder, pady=3
        ).pack(side=tk.LEFT, padx=(6, 0))

        # Hiển thị số dòng / tài khoản hợp lệ
        self.line_count_lbl = tk.Label(
            imp_row, text="0 dòng",
            font=small9, bg=self.COLOR_BG, fg=self.COLOR_MUTED
        )
        self.line_count_lbl.pack(side=tk.RIGHT, padx=4)

        self.acc_text = tk.Text(
            left, width=52, height=14,
            font=mono10, bg=self.COLOR_PANEL, fg=self.COLOR_TEXT,
            insertbackground=self.COLOR_TEXT, relief=tk.FLAT, padx=8, pady=6,
            wrap=tk.NONE
        )
        self.acc_text.pack(fill=tk.BOTH, expand=True)

        # Placeholder mờ: chỉ hiện gợi ý khi ô trống, tự ẩn khi gõ/dán.
        self._acc_placeholder = (
            "Dán account hoặc bấm '📂 Nhập từ file'.  Cột: mail | pass | recovery | 2fa\n"
            "(2fa để trống nếu account không bật 2FA). Ví dụ:\n"
            "kenh1@gmail.com | MatKhau123 | backup1@gmail.com | JBSWY3DPEHPK3PXP\n"
            "kenh2@gmail.com | MatKhau456 | backup2@gmail.com |"
        )
        self._acc_ph_on = False
        self._show_acc_placeholder()
        self.acc_text.bind("<FocusIn>", self._acc_focus_in)
        self.acc_text.bind("<FocusOut>", self._acc_focus_out)
        # DEBOUNCE: với list account lớn, đếm dòng mỗi phím gây LAG → chỉ đếm sau khi
        # ngừng gõ ~350ms (huỷ lịch cũ mỗi lần gõ).
        self.acc_text.bind("<KeyRelease>", lambda e: self._debounce_line_count())
        self.acc_text.bind("<<Paste>>", lambda e: self._debounce_line_count(400))

        # ── Right: Profile + options panel ──
        right = tk.Frame(body, bg=self.COLOR_BG, width=250)
        right.pack(side=tk.RIGHT, fill=tk.Y, padx=(6, 0))
        right.pack_propagate(False)

        # (ĐÃ BỎ khối 'GPM Profiles' + nút 'Tải lại profiles' khỏi giao diện — không cần
        #  hiển thị danh sách profile. Vẫn nạp self.profiles ngầm để so khớp tên khi cần.)
        self.profile_lb = None

        # API Token (tự nạp danh sách key)
        _tokhdr = tk.Frame(right, bg=self.COLOR_BG)
        _tokhdr.pack(fill=tk.X, anchor=tk.W, pady=(8, 0))
        tk.Label(_tokhdr, text="API Token (kiotproxy):",
                 font=norm10, bg=self.COLOR_BG, fg=self.COLOR_MUTED).pack(side=tk.LEFT)
        self.token_show_var = tk.BooleanVar(value=False)
        tk.Checkbutton(_tokhdr, text="👁 hiện", variable=self.token_show_var,
                       command=self._toggle_token_show, font=small9,
                       bg=self.COLOR_BG, fg=self.COLOR_MUTED, selectcolor=self.COLOR_PANEL,
                       activebackground=self.COLOR_BG, relief=tk.FLAT).pack(side=tk.RIGHT)
        self.token_var = tk.StringVar(value=os.environ.get("KIOT_API_TOKEN", ""))
        self.token_entry = tk.Entry(right, textvariable=self.token_var, show="•",
                 font=mono9, bg=self.COLOR_PANEL, fg=self.COLOR_TEXT,
                 insertbackground=self.COLOR_TEXT, relief=tk.FLAT)
        self.token_entry.pack(fill=tk.X, pady=(2, 2))
        # TỰ LƯU token NGAY khi gõ (ghi đè config.env) → mở lại vẫn giữ token của MÁY NÀY.
        self.token_var.trace_add("write", lambda *a: self._persist_token(self.token_var.get().strip()))
        tk.Button(
            right, text="⬇ Tải keys từ tài khoản",
            font=small9, bg=self.COLOR_PANEL, fg=self.COLOR_TEXT,
            relief=tk.FLAT, cursor="hand2", activebackground=self.COLOR_ACCENT,
            command=self._load_keys_from_token, pady=4
        ).pack(fill=tk.X)

        # ── Google Sheet (đọc account + ghi kết quả online) ──
        _gs = self._gsheet_load_cfg()
        tk.Label(right, text="\nGoogle Sheet (Apps Script /exec URL):",
                 font=norm10, bg=self.COLOR_BG, fg=self.COLOR_MUTED).pack(anchor=tk.W)
        self.gsheet_url_var = tk.StringVar(value=_gs.get("url", ""))
        tk.Entry(right, textvariable=self.gsheet_url_var,
                 font=mono9, bg=self.COLOR_PANEL, fg=self.COLOR_TEXT,
                 insertbackground=self.COLOR_TEXT, relief=tk.FLAT).pack(fill=tk.X, pady=(2, 2))
        self.gsheet_secret_var = tk.StringVar(value=_gs.get("secret", "hapu-2fa-secret"))
        tk.Entry(right, textvariable=self.gsheet_secret_var, show="•",
                 font=mono9, bg=self.COLOR_PANEL, fg=self.COLOR_TEXT,
                 insertbackground=self.COLOR_TEXT, relief=tk.FLAT).pack(fill=tk.X, pady=(0, 2))
        # TÊN MÁY — quyết định cặp tab riêng của máy này: NHAP_<tên> (đọc) / KETQUA_<tên> (ghi).
        # Mỗi máy đặt tên khác nhau → chạy chung 1 link Sheet mà không đụng nhau.
        tk.Label(right, text="Tên máy (tab riêng: NHAP_<tên> / KETQUA_<tên>):",
                 font=small9, bg=self.COLOR_BG, fg=self.COLOR_MUTED).pack(anchor=tk.W)
        self.machine_var = tk.StringVar(value=_gs.get("machine", ""))
        tk.Entry(right, textvariable=self.machine_var,
                 font=mono9, bg=self.COLOR_PANEL, fg=self.COLOR_TEXT,
                 insertbackground=self.COLOR_TEXT, relief=tk.FLAT).pack(fill=tk.X, pady=(2, 2))
        # TỰ LƯU tên máy NGAY khi gõ (khỏi phải chạy/test mới lưu) → lần sau mở tool trên
        # máy đó vẫn còn nguyên tên, không phải nhập lại.
        self.machine_var.trace_add("write", lambda *a: self._gsheet_save_cfg())
        self.gsheet_on_var = tk.BooleanVar(value=_gs.get("enabled", False))
        tk.Checkbutton(right, text="Dùng Google Sheet (đọc/ghi online)",
                       variable=self.gsheet_on_var, font=small9,
                       bg=self.COLOR_BG, fg="#66ccff", selectcolor=self.COLOR_PANEL,
                       activebackground=self.COLOR_BG, relief=tk.FLAT).pack(anchor=tk.W)
        self.gsheet_nolocal_var = tk.BooleanVar(value=_gs.get("nolocal", True))
        tk.Checkbutton(right, text="Không tải Excel local (chỉ ghi Sheet)",
                       variable=self.gsheet_nolocal_var, font=small9,
                       bg=self.COLOR_BG, fg=self.COLOR_TEXT, selectcolor=self.COLOR_PANEL,
                       activebackground=self.COLOR_BG, relief=tk.FLAT).pack(anchor=tk.W)
        _gsrow = tk.Frame(right, bg=self.COLOR_BG)
        _gsrow.pack(fill=tk.X, pady=(2, 0))
        self._gsheet_load_btn = tk.Button(
                  _gsrow, text="📥 Nạp account từ Sheet", font=small9,
                  bg=self.COLOR_PANEL, fg=self.COLOR_TEXT, relief=tk.FLAT, cursor="hand2",
                  activebackground=self.COLOR_ACCENT, command=self._gsheet_load_accounts,
                  pady=4)
        self._gsheet_load_btn.pack(side=tk.LEFT, fill=tk.X, expand=True)
        tk.Button(_gsrow, text="🔗 Test", font=small9,
                  bg=self.COLOR_PANEL, fg=self.COLOR_TEXT, relief=tk.FLAT, cursor="hand2",
                  activebackground=self.COLOR_ACCENT, command=self._gsheet_test,
                  pady=4).pack(side=tk.LEFT, padx=(4, 0))

        # Region
        tk.Label(right, text="\nVùng proxy (region):",
                 font=norm10, bg=self.COLOR_BG, fg=self.COLOR_MUTED).pack(anchor=tk.W)
        self.region_var = tk.StringVar(value="random")
        ttk.Combobox(right, textvariable=self.region_var, values=REGIONS,
                     state="readonly", width=12).pack(anchor=tk.W, pady=(2, 4))

        # Matching mode
        tk.Label(right, text="Match profile theo:",
                 font=norm10, bg=self.COLOR_BG, fg=self.COLOR_MUTED).pack(anchor=tk.W)
        self.match_var = tk.StringVar(value="auto")
        for val, text in [("auto", "Tự động (theo tên email)"), ("pick", "Chọn thủ công ↑")]:
            tk.Radiobutton(
                right, text=text, variable=self.match_var, value=val,
                font=small9, bg=self.COLOR_BG, fg=self.COLOR_TEXT,
                selectcolor=self.COLOR_PANEL, activebackground=self.COLOR_BG,
                relief=tk.FLAT
            ).pack(anchor=tk.W)

        self.auto_split_var = tk.BooleanVar(value=True)
        tk.Checkbutton(
            right, text="Tự chia account đều cho các key",
            variable=self.auto_split_var, font=small9,
            bg=self.COLOR_BG, fg=self.COLOR_TEXT,
            selectcolor=self.COLOR_PANEL, activebackground=self.COLOR_BG,
            relief=tk.FLAT
        ).pack(anchor=tk.W, pady=(6, 0))

        # Số luồng chạy song song
        thr_row = tk.Frame(right, bg=self.COLOR_BG)
        thr_row.pack(anchor=tk.W, fill=tk.X, pady=(6, 0))
        tk.Label(thr_row, text="Số luồng song song:", font=small9,
                 bg=self.COLOR_BG, fg=self.COLOR_TEXT).pack(side=tk.LEFT)
        self.concurrency_var = tk.IntVar(
            value=int(self._gsheet_load_cfg().get("threads", 1) or 1))
        tk.Spinbox(thr_row, from_=1, to=20, width=4, textvariable=self.concurrency_var,
                   font=small9, justify=tk.CENTER).pack(side=tk.LEFT, padx=(6, 0))

        self.create_new_var = tk.BooleanVar(value=True)
        tk.Checkbutton(
            right, text="Tạo profile mới nếu không có",
            variable=self.create_new_var, font=small9,
            bg=self.COLOR_BG, fg=self.COLOR_TEXT,
            selectcolor=self.COLOR_PANEL, activebackground=self.COLOR_BG,
            relief=tk.FLAT
        ).pack(anchor=tk.W, pady=(6, 0))

        self.keep_open_var = tk.BooleanVar(value=False)
        tk.Checkbutton(
            right, text="Giữ browser mở sau login",
            variable=self.keep_open_var, font=small9,
            bg=self.COLOR_BG, fg=self.COLOR_TEXT,
            selectcolor=self.COLOR_PANEL, activebackground=self.COLOR_BG,
            relief=tk.FLAT
        ).pack(anchor=tk.W)

        self.test_proxy_var = tk.BooleanVar(value=True)
        tk.Checkbutton(
            right, text="Test proxy trước khi login",
            variable=self.test_proxy_var, font=small9,
            bg=self.COLOR_BG, fg=self.COLOR_TEXT,
            selectcolor=self.COLOR_PANEL, activebackground=self.COLOR_BG,
            relief=tk.FLAT
        ).pack(anchor=tk.W)

        # CHẾ ĐỘ TEST: dừng khi gặp lỗi + nhớ tiến độ (chạy lại bỏ qua account đã xong).
        # Chỉ dùng khi test/sửa tool. Chạy bình thường TẮT (giữ nguyên như cũ).
        self.test_mode_var = tk.BooleanVar(
            value=bool(self._gsheet_load_cfg().get("test", False)))
        tk.Checkbutton(
            right, text="🧪 Chế độ TEST (dừng khi lỗi + nhớ tiến độ)",
            variable=self.test_mode_var, font=small9,
            bg=self.COLOR_BG, fg="#ffcc66",
            selectcolor=self.COLOR_PANEL, activebackground=self.COLOR_BG,
            relief=tk.FLAT
        ).pack(anchor=tk.W)

        tk.Button(
            right, text="🔎 Kiểm tra key (lấy IP hiện tại)",
            font=small9, bg=self.COLOR_PANEL, fg=self.COLOR_TEXT,
            relief=tk.FLAT, cursor="hand2", activebackground=self.COLOR_ACCENT,
            command=self._test_keys, pady=4
        ).pack(fill=tk.X, pady=(8, 0))

        # ── Buttons (nằm dưới tab1, luôn hiện) ──
        btn_row = tk.Frame(tab1, bg=self.COLOR_BG)
        btn_row.pack(fill=tk.X, padx=12, pady=(0, 8))

        self.start_btn = tk.Button(
            btn_row, text="▶  ĐĂNG NHẬP",
            font=tkfont.Font(family="Segoe UI", size=11, weight="bold"),
            bg=self.COLOR_GREEN, fg="white", relief=tk.FLAT, cursor="hand2",
            padx=24, pady=8, activebackground="#16a34a",
            command=self._start
        )
        self.start_btn.pack(side=tk.LEFT, padx=(0, 8))

        tk.Button(
            btn_row, text="⬛  Dừng",
            font=tkfont.Font(family="Segoe UI", size=11),
            bg=self.COLOR_RED, fg="white", relief=tk.FLAT, cursor="hand2",
            padx=18, pady=8, activebackground="#b91c1c",
            command=self._stop
        ).pack(side=tk.LEFT)

        # Chỉnh số luồng ngay cạnh nút (dễ thấy) ────────────────
        thr_box = tk.Frame(btn_row, bg=self.COLOR_BG)
        thr_box.pack(side=tk.LEFT, padx=(14, 0))
        tk.Label(thr_box, text="Số luồng:", bg=self.COLOR_BG, fg="white",
                 font=tkfont.Font(family="Segoe UI", size=11, weight="bold")
                 ).pack(side=tk.LEFT, padx=(0, 4))
        tk.Spinbox(
            thr_box, from_=1, to=20, width=4, textvariable=self.concurrency_var,
            font=tkfont.Font(family="Segoe UI", size=12, weight="bold"),
            justify="center", buttonbackground=self.COLOR_PANEL,
        ).pack(side=tk.LEFT)

        tk.Button(
            btn_row, text="💾  Xuất kết quả (CSV)",
            font=tkfont.Font(family="Segoe UI", size=11),
            bg=self.COLOR_PANEL, fg="white", relief=tk.FLAT, cursor="hand2",
            padx=18, pady=8, activebackground=self.COLOR_ACCENT,
            command=self._save_results
        ).pack(side=tk.LEFT, padx=(8, 0))

        self.status_lbl = tk.Label(
            btn_row, text="Sẵn sàng",
            font=norm10, bg=self.COLOR_BG, fg=self.COLOR_MUTED
        )
        self.status_lbl.pack(side=tk.RIGHT, padx=8)

        # Hiển thị tiến độ: X / Y tài khoản
        self.progress_lbl = tk.Label(
            btn_row, text="",
            font=tkfont.Font(family="Segoe UI", size=10, weight="bold"),
            bg=self.COLOR_BG, fg=self.COLOR_ACCENT
        )
        self.progress_lbl.pack(side=tk.RIGHT, padx=12)

        # ── Tab 2: Kết quả & Log ─────────────────────────────
        tab2 = tk.Frame(self.main_nb, bg=self.COLOR_BG)
        self.main_nb.add(tab2, text="▶  Kết quả & Log")

        # Thanh nút nhỏ trong tab2 (Dừng + Xuất + tiến độ)
        tab2_bar = tk.Frame(tab2, bg=self.COLOR_BG)
        tab2_bar.pack(fill=tk.X, padx=12, pady=(6, 2))
        tk.Button(tab2_bar, text="⬛  Dừng",
                  font=tkfont.Font(family="Segoe UI", size=10),
                  bg=self.COLOR_RED, fg="white", relief=tk.FLAT, cursor="hand2",
                  padx=14, pady=6, activebackground="#b91c1c",
                  command=self._stop).pack(side=tk.LEFT)
        tk.Button(tab2_bar, text="💾  Xuất kết quả",
                  font=tkfont.Font(family="Segoe UI", size=10),
                  bg=self.COLOR_PANEL, fg="white", relief=tk.FLAT, cursor="hand2",
                  padx=14, pady=6, activebackground=self.COLOR_ACCENT,
                  command=self._save_results).pack(side=tk.LEFT, padx=(6, 0))
        self.progress_lbl2 = tk.Label(tab2_bar, text="",
                  font=tkfont.Font(family="Segoe UI", size=10, weight="bold"),
                  bg=self.COLOR_BG, fg=self.COLOR_ACCENT)
        self.progress_lbl2.pack(side=tk.RIGHT, padx=12)
        self.status_lbl2 = tk.Label(tab2_bar, text="Sẵn sàng",
                  font=tkfont.Font(family="Segoe UI", size=10),
                  bg=self.COLOR_BG, fg=self.COLOR_MUTED)
        self.status_lbl2.pack(side=tk.RIGHT, padx=8)

        nb = ttk.Notebook(tab2)
        nb.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 8))

        # Tab: Bảng kết quả (theo dõi từng luồng + trạng thái)
        tbl = tk.Frame(nb, bg=self.COLOR_BG)
        nb.add(tbl, text="📊 Bảng kết quả")
        cols = ("stt", "luong", "email", "matkhau", "recovery", "twofa",
                "ip", "trangthai", "chitiet")
        heads = {"stt": "#", "luong": "Luồng", "email": "Email", "matkhau": "Mật khẩu",
                 "recovery": "Recovery mail", "twofa": "2FA secret",
                 "ip": "IP proxy", "trangthai": "Trạng thái", "chitiet": "Chi tiết"}
        widths = {"stt": 36, "luong": 45, "email": 175, "matkhau": 110, "recovery": 175,
                  "twofa": 120, "ip": 175, "trangthai": 150, "chitiet": 230}
        self.tree = ttk.Treeview(tbl, columns=cols, show="headings", height=12)
        for c in cols:
            self.tree.heading(c, text=heads[c])
            self.tree.column(c, width=widths[c], anchor=tk.W)
        vsb = ttk.Scrollbar(tbl, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.tree.column("stt", anchor=tk.CENTER)
        self.tree.tag_configure("ok", foreground="#16a34a")
        self.tree.tag_configure("err", foreground="#dc2626")
        self.tree.tag_configure("warn", foreground="#d97706")
        self.tree.tag_configure("run", foreground="#2563eb")
        self._tree_rows: dict = {}

        # Tab: Log
        lf = tk.Frame(nb, bg=self.COLOR_BG)
        nb.add(lf, text="📜 Log")
        self.log_box = scrolledtext.ScrolledText(
            lf, height=12, font=mono9, bg="#0f0f1a", fg=self.COLOR_TEXT,
            state=tk.DISABLED, relief=tk.FLAT, padx=8, pady=6
        )
        self.log_box.pack(fill=tk.BOTH, expand=True)
        self.log_box.tag_config("ok", foreground=self.COLOR_GREEN)
        self.log_box.tag_config("err", foreground=self.COLOR_RED)
        self.log_box.tag_config("warn", foreground=self.COLOR_YELLOW)
        self.log_box.tag_config("muted", foreground=self.COLOR_MUTED)

    # ── Bảng kết quả (Treeview) ────────────────────────────
    def _tree_clear(self):
        def _do():
            for iid in self.tree.get_children():
                self.tree.delete(iid)
            self._tree_rows = {}
            self._stt_counter = 0
        self.root.after(0, _do)

    def _tree_upsert(self, email, luong=None, ip=None, status=None, detail=None,
                     tag="", matkhau=None, recovery=None, twofa=None):
        def _do():
            iid = self._tree_rows.get(email)
            if iid and self.tree.exists(iid):
                cur = list(self.tree.item(iid, "values"))
                # col order: stt(0) luong(1) email(2) matkhau(3) recovery(4) twofa(5) ip(6) status(7) detail(8)
                if luong is not None:
                    cur[1] = luong
                if matkhau is not None:
                    cur[3] = matkhau
                if recovery is not None:
                    cur[4] = recovery
                if twofa is not None:
                    cur[5] = twofa
                if ip:
                    cur[6] = ip
                if status is not None:
                    cur[7] = status
                if detail is not None:
                    cur[8] = (detail or "")[:200]
                self.tree.item(iid, values=cur, tags=((tag,) if tag else self.tree.item(iid, "tags")))
            else:
                self._stt_counter += 1
                iid = self.tree.insert(
                    "", tk.END,
                    values=(self._stt_counter, luong or "", email, matkhau or "",
                            recovery or "", twofa or "", ip or "",
                            status or "", (detail or "")[:200]),
                    tags=((tag,) if tag else ()))
                self._tree_rows[email] = iid
            self.tree.see(iid)
        self.root.after(0, _do)

    # ── Profile helpers ────────────────────────────────────
    def _load_profiles(self):
        try:
            self.profiles = gpm_list()
            if getattr(self, "profile_lb", None) is not None:
                self.profile_lb.delete(0, tk.END)
                for p in self.profiles:
                    name = p.get("name") or p.get("profileName") or "?"
                    self.profile_lb.insert(tk.END, f"{name}")
            self._log(f"✓ Tải {len(self.profiles)} profiles GPM", "ok")
        except Exception as e:
            self._log(f"✗ Không kết nối GPM-Login: {e}\n  → Hãy mở GPM-Login trước!", "err")

    def _find_profile(self, email: str):
        """Trả về (profile_id, profile_name) hoặc (None, None)."""
        if self.match_var.get() == "pick":
            if getattr(self, "profile_lb", None) is None:
                return None, None
            sel = self.profile_lb.curselection()
            if sel:
                p = self.profiles[sel[0]]
                return p.get("id") or p.get("profileId"), p.get("name")
            return None, None
        elow = email.lower()
        eshort = elow.split("@")[0]
        for p in self.profiles:
            name = (p.get("name") or "").lower()
            if elow in name or eshort in name:
                return p.get("id") or p.get("profileId"), p.get("name")
        return None, None

    # ── Logging ────────────────────────────────────────────
    def _log(self, msg: str, tag: str = ""):
        # Ghi thêm ra stdout (run_log.txt) để tiện chẩn đoán ngoài GUI
        try:
            print(msg.encode("ascii", "replace").decode("ascii"), flush=True)
        except Exception:
            pass
        def _do():
            self.log_box.config(state=tk.NORMAL)
            self.log_box.insert(tk.END, msg + "\n", tag)
            self.log_box.see(tk.END)
            self.log_box.config(state=tk.DISABLED)
        self.root.after(0, _do)

    def _set_status(self, msg: str):
        def _do():
            self.status_lbl.config(text=msg)
            self.status_lbl2.config(text=msg)
        self.root.after(0, _do)

    def _debounce_line_count(self, delay: int = 350):
        """Lịch đếm dòng sau `delay`ms, HUỶ lịch cũ → không tính lại mỗi phím (đỡ lag)."""
        try:
            if getattr(self, "_lc_after", None):
                self.root.after_cancel(self._lc_after)
        except Exception:
            pass
        self._lc_after = self.root.after(delay, self._update_line_count)

    def _update_line_count(self):
        """Đếm số dòng tài khoản hợp lệ trong ô nhập và cập nhật nhãn."""
        if self._acc_ph_on:
            self.line_count_lbl.config(text="0 dòng")
            return
        raw = self.acc_text.get("1.0", tk.END)
        count = sum(
            1 for ln in raw.splitlines()
            if ln.strip() and not ln.strip().startswith("#")
        )
        txt = f"{count} dòng" if count != 1 else "1 dòng"
        self.line_count_lbl.config(
            text=txt,
            fg=self.COLOR_GREEN if count > 0 else self.COLOR_MUTED
        )

    # ── Placeholder mờ cho ô nhập account ─────────────────
    def _show_acc_placeholder(self):
        self.acc_text.delete("1.0", tk.END)
        self.acc_text.insert("1.0", self._acc_placeholder)
        self.acc_text.tag_add("ph", "1.0", tk.END)
        self.acc_text.tag_config("ph", foreground="#6b7280")  # xám mờ
        self._acc_ph_on = True

    def _clear_acc_placeholder(self):
        if self._acc_ph_on:
            self.acc_text.delete("1.0", tk.END)
            self.acc_text.tag_remove("ph", "1.0", tk.END)
            self.acc_text.config(fg=self.COLOR_TEXT)
            self._acc_ph_on = False

    def _acc_focus_in(self, _e=None):
        self._clear_acc_placeholder()

    def _acc_focus_out(self, _e=None):
        if not self.acc_text.get("1.0", tk.END).strip():
            self._show_acc_placeholder()

    def _acc_get_text(self) -> str:
        """Nội dung thật của ô (rỗng nếu đang hiện placeholder)."""
        if self._acc_ph_on:
            return ""
        return self.acc_text.get("1.0", tk.END)

    def _parse_input(self):
        """
        Đọc ô nhập. Cột mỗi dòng: gmail | pass | recovery | 2fa_secret.
        Key lấy từ dòng 'KEY: xxx' (nếu có). Account không nằm dưới KEY nào -> keyless
        (sẽ được tự chia đều cho các key đã tải từ token).
        Trả về (groups: OrderedDict[key -> list[acc]], keyless: list[acc]).
        """
        groups: "OrderedDict[str, list[dict]]" = OrderedDict()
        keyless: list[dict] = []
        current_key = None
        for raw in self._acc_get_text().splitlines():
            line = raw.strip()
            if not line:
                continue
            head = line.lstrip("#").strip()
            if head.lower().startswith("key:") or head.lower().startswith("key ="):
                current_key = head.split(":", 1)[-1].split("=", 1)[-1].strip()
                if current_key:
                    groups.setdefault(current_key, [])
                continue
            if line.startswith("#"):
                continue
            sep = "|" if "|" in line else ","
            parts = [p.strip() for p in line.split(sep)]
            if len(parts) < 2 or not parts[0]:
                continue
            # Bỏ dòng tiêu đề nếu có
            if parts[0].lower() in ("gmail", "email", "mail", "tài khoản",
                                     "account", "username", "user", "tên"):
                continue
            # Mail có thể là 'name@gmail.com' hoặc chỉ 'name' -> tự thêm @gmail.com
            mail = parts[0]
            if "@" not in mail:
                mail = f"{mail}@gmail.com"

            # ── Tự nhận dạng định dạng cột ──────────────────────
            # Dạng 1: mail|pass|recovery|2fa   (4 cột, cột 3 có @)
            # Dạng 2: mail|pass|recovery|       (4 cột, cột 4 rỗng)
            # Dạng 3: mail|pass|2fa             (3 cột, cột 3 KHÔNG có @)
            pw       = parts[1] if len(parts) > 1 else ""
            recovery = ""
            totp     = ""
            if len(parts) >= 4:
                # 4 cột: cột 3 là recovery, cột 4 là 2FA
                recovery = parts[2]
                totp     = parts[3]
                # NHƯNG nếu cột 4 (2fa) RỖNG và cột 3 KHÔNG có '@' → cột 3 thực ra là
                # 2FA secret (lô mail có cột 3 = 2FA HOẶC recovery tuỳ dòng).
                if not totp.strip() and recovery and "@" not in recovery:
                    totp = recovery
                    recovery = ""
            elif len(parts) == 3:
                f3 = parts[2]
                if "@" in f3:
                    # Có @ → là recovery mail, không có 2FA
                    recovery = f3
                else:
                    # Không có @ → là 2FA secret
                    totp = f3

            # Cột E (thứ 5) = MAIL QUẢN TRỊ để add owner kênh thương hiệu (nhiệm vụ Add QT TH).
            # Không ảnh hưởng cột 1-4; dòng cũ 4 cột → owner rỗng.
            owner = parts[4].strip() if len(parts) > 4 else ""
            acc = {
                "email":    mail,
                "password": pw,
                "recovery": recovery,
                "totp":     totp,
                "owner":    owner,
            }
            if current_key:
                groups.setdefault(current_key, []).append(acc)
            else:
                keyless.append(acc)
        groups = OrderedDict((k, v) for k, v in groups.items() if v)
        return groups, keyless

    def _parse_groups(self):
        g, _ = self._parse_input()
        return g

    def _build_run_plan(self):
        """Gộp group tường minh (KEY:) + tự chia account keyless đều cho các key đã tải.
        Trả về (groups, err_msg)."""
        groups, keyless = self._parse_input()
        groups = OrderedDict((k, list(v)) for k, v in groups.items())
        if keyless:
            keys = list(self.available_keys) or list(groups.keys())
            if not keys:
                return None, ("Chưa có key. Bấm '⬇ Tải keys từ tài khoản' trước "
                              "(hoặc thêm dòng 'KEY: <key>').")
            if self.auto_split_var.get() or not groups:
                for i, acc in enumerate(keyless):
                    k = keys[i % len(keys)]
                    groups.setdefault(k, []).append(acc)
        if not groups:
            return None, "Không có tài khoản hợp lệ trong ô nhập."
        return groups, None

    # ── Nhập tài khoản từ file ─────────────────────────────
    def _read_accounts_file(self, path: str) -> list:
        """
        Đọc file tài khoản, trả về list dòng theo định dạng ô nhập.
        - .txt: giữ nguyên (hỗ trợ cả dòng 'KEY: xxx' và account).
        - .csv/.xlsx/.xls: cột Gmail | Mật khẩu | Mail khôi phục | [Key].
          Dòng có cột Key thì xuất 4 cột; không có thì 3 cột.
        """
        ext = os.path.splitext(path)[1].lower()
        if ext == ".txt":
            with open(path, encoding="utf-8", errors="replace") as f:
                return [ln.rstrip("\n") for ln in f if ln.strip()]

        rows = []
        if ext == ".csv":
            import csv
            with open(path, encoding="utf-8-sig", errors="replace", newline="") as f:
                sample = f.read(4096); f.seek(0)
                delim = "|" if "|" in sample else ","
                for r in csv.reader(f, delimiter=delim):
                    rows.append([("" if c is None else str(c)).strip() for c in r])
        elif ext in (".xlsx", ".xls"):
            try:
                import openpyxl
            except ImportError:
                import subprocess as sp, sys
                sp.run([sys.executable, "-m", "pip", "install", "openpyxl",
                        "-q", "--break-system-packages"])
                import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            for r in ws.iter_rows(values_only=True):
                rows.append([("" if c is None else str(c)).strip() for c in r])
        else:
            raise ValueError(f"Định dạng không hỗ trợ: {ext} (chỉ nhận xlsx/csv/txt)")

        lines = []
        header_words = ("gmail", "email", "mail", "tài khoản", "account",
                        "username", "user", "tên", "mật khẩu", "password")
        for cols in rows:
            if not cols or not cols[0]:
                continue
            gmail = cols[0]
            # Bỏ dòng tiêu đề (nếu có) — dựa trên từ khóa, KHÔNG dựa vào '@'
            # để chấp nhận mail dạng tên trần (không có @gmail.com).
            if gmail.strip().lower() in header_words:
                continue
            pw  = cols[1] if len(cols) > 1 else ""
            c2  = cols[2].strip() if len(cols) > 2 else ""
            c3  = cols[3].strip() if len(cols) > 3 else ""
            # Tự nhận dạng: nếu có cột 4 thì c2=recovery, c3=2FA
            # Nếu chỉ có cột 3: @ → recovery, còn lại → 2FA
            if c3:
                rec, totp = c2, c3
            elif c2 and "@" in c2:
                rec, totp = c2, ""
            else:
                rec, totp = "", c2
            line = f"{gmail}|{pw}|{rec}|{totp}"
            lines.append(line)
        return lines

    def _import_from_file(self):
        path = filedialog.askopenfilename(
            title="Chọn file tài khoản",
            filetypes=[
                ("Tài khoản (xlsx/csv/txt)", "*.xlsx *.xls *.csv *.txt"),
                ("Excel", "*.xlsx *.xls"),
                ("CSV", "*.csv"),
                ("Text", "*.txt"),
                ("Tất cả", "*.*"),
            ],
        )
        if not path:
            return
        try:
            lines = self._read_accounts_file(path)
        except Exception as e:
            messagebox.showerror("Lỗi đọc file", str(e))
            self._log(f"✗ Lỗi đọc file: {e}", "err")
            return
        if not lines:
            messagebox.showwarning("Trống", "Không đọc được dòng tài khoản nào trong file.")
            return
        # Xóa sạch ô nhập (kể cả placeholder mờ) rồi nạp đúng file
        self._clear_acc_placeholder()
        self.acc_text.delete("1.0", tk.END)
        self.acc_text.insert("1.0", "\n".join(lines) + "\n")
        self.acc_text.config(fg=self.COLOR_TEXT)
        self._acc_ph_on = False
        self._update_line_count()
        has_2fa = any(l.count("|") >= 3 for l in lines)
        note = "  (có cột 2FA ✓)" if has_2fa else ""
        self._log(f"📂 Đã nhập {len(lines)} tài khoản từ {os.path.basename(path)}{note}. "
                  f"Nhớ bấm '⬇ Tải keys' để tool tự chia key.", "ok")

    def _toggle_token_show(self):
        """Hiện/ẩn API token."""
        try:
            self.token_entry.config(show="" if self.token_show_var.get() else "•")
        except Exception:
            pass

    # ── Tự nạp danh sách key qua API Token ─────────────────
    def _persist_token(self, token: str):
        """Lưu/CẬP NHẬT KIOT_API_TOKEN vào config.env (GHI ĐÈ cái cũ) → mỗi máy giữ token
        riêng mình gõ, mở lại vẫn còn, KHÔNG bị nhảy về token hệ thống ban đầu."""
        if not token:
            return
        try:
            import re as _re
            cfg = Path(__file__).parent / "config.env"
            content = cfg.read_text(encoding="utf-8") if cfg.exists() else ""
            line = f"KIOT_API_TOKEN={token}"
            if _re.search(r'(?m)^\s*KIOT_API_TOKEN=.*$', content):
                content = _re.sub(r'(?m)^\s*KIOT_API_TOKEN=.*$', line, content)   # ghi đè dòng cũ
            else:
                if content and not content.endswith("\n"):
                    content += "\n"
                content += f"# API Token kiotproxy (dùng cho tool đăng nhập)\n{line}\n"
            cfg.write_text(content, encoding="utf-8")
        except Exception:
            pass

    @staticmethod
    def _exp_timestamp(exp):
        """Đổi expiration (epoch giây/mili giây hoặc chuỗi ISO) -> timestamp giây, hoặc None."""
        if exp in (None, ""):
            return None
        if isinstance(exp, (int, float)):
            return exp / 1000.0 if exp > 1e12 else float(exp)
        try:
            import datetime
            s = str(exp).strip().replace("Z", "+00:00")
            dt = datetime.datetime.fromisoformat(s)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=datetime.timezone.utc)
            return dt.timestamp()
        except Exception:
            return None

    @staticmethod
    def _exp_display(exp) -> str:
        """Chuỗi hạn dùng để hiển thị (yyyy-mm-dd)."""
        ts = KiotLoginApp._exp_timestamp(exp)
        if ts is None:
            return ""
        import datetime
        try:
            return datetime.datetime.fromtimestamp(ts).strftime("%Y-%m-%d")
        except Exception:
            return ""

    @staticmethod
    def _key_usable(k: dict) -> bool:
        """Key dùng được = status ACTIVE và chưa hết hạn (expirationAt > bây giờ)."""
        if (k.get("status") or "").upper() != "ACTIVE":
            return False
        ts = KiotLoginApp._exp_timestamp(k.get("expiration"))
        if ts is None:
            return True   # không có/không parse được hạn -> vẫn cho dùng
        return ts > time.time()

    def _load_keys_from_token(self):
        token = self.token_var.get().strip()
        if not token:
            messagebox.showwarning("Thiếu token", "Dán API Token kiotproxy vào ô 'API Token'.")
            return

        def run():
            try:
                info = self.kiot.user_info(token)
                if info:
                    self._log(f"👤 Tài khoản: {info.get('username','?')} | "
                              f"Số dư: {info.get('balance','?')} | maxKey: {info.get('maxKey','?')}", "ok")
            except Exception as e:
                self._log(f"⚠ Không lấy được thông tin tài khoản: {e}", "warn")

            try:
                keys = self.kiot.list_keys(token)
            except KiotError as e:
                self._log(f"✗ Token sai hoặc lỗi API: {e}", "err")
                return
            except Exception as e:
                self._log(f"✗ Lỗi kết nối kiotproxy: {e}", "err")
                return

            active = [k for k in keys if self._key_usable(k)]
            expired = len(keys) - len(active)
            self.available_keys = [k.get("value", "") for k in active if k.get("value")]
            self._persist_token(token)
            self._log(f"⬇ Nhận {len(keys)} key → dùng {len(self.available_keys)} key CÒN HẠN "
                      f"({expired} key hết hạn/không active bị bỏ). Tool sẽ TỰ CHIA account cho các key này.",
                      "ok")
            for k in active[:40]:
                exp = self._exp_display(k.get("expiration"))
                desc = k.get("description", "") or ""
                self._log(f"   • {str(k.get('value',''))[:10]}…  {('['+desc+'] ') if desc else ''}"
                          f"{('HSD '+exp) if exp else ''}", "muted")
            self._log("→ Dán/nhập account (mail | pass | recovery | 2fa) rồi bấm BẮT ĐẦU.", "ok")

        threading.Thread(target=run, daemon=True).start()

    # ── Kiểm tra key ───────────────────────────────────────
    def _test_keys(self):
        keys = list(self.available_keys) or list(self._parse_groups().keys())
        if not keys:
            messagebox.showwarning("Thiếu key",
                                   "Bấm '⬇ Tải keys từ tài khoản' trước, hoặc thêm dòng 'KEY:'.")
            return

        def run():
            for key in keys:
                try:
                    info = self.kiot.get_current(key)
                    self._log(f"🔎 {key[:8]}… IP hiện tại: {info['host']}:{info['http_port']} "
                              f"@ {info['location']} (ttc={info['ttc']}s)", "ok")
                except KiotError as e:
                    self._log(f"🔎 {key[:8]}… chưa có proxy/hoặc lỗi: {e} — thử lấy mới…", "warn")
                    try:
                        info = self.kiot.get_new(key, self.region_var.get())
                        self._log(f"    → IP mới: {info['host']}:{info['http_port']} @ {info['location']}", "ok")
                    except Exception as e2:
                        self._log(f"    ✗ Lỗi: {e2}", "err")
                except Exception as e:
                    self._log(f"🔎 {key[:8]}… lỗi kết nối: {e}", "err")

        threading.Thread(target=run, daemon=True).start()

    # ── Cập nhật nút BẮT ĐẦU theo tác vụ đã chọn ──────────
    def _refresh_start_btn(self):
        selected = self._selected_ordered()   # theo ĐÚNG thứ tự click
        if not selected:
            self.start_btn.config(text="▶  BẮT ĐẦU")
            return
        _names = {
            "login":     "ĐĂNG NHẬP",
            "changepw":  "ĐỔI MẬT KHẨU",
            "change2fa": "ĐỔI 2FA",
            "create2fa": "TẠO 2FA",
            "channel":   "TẠO KÊNH",
            "getlink":   "LẤY LINK KÊNH",
            "banner":    "THAY ẢNH BÌA",
            "rmadmin":   "THOÁT QUẢN TRỊ",
            "addqtth":   "ADD QT TH",
            "addqtonly": "ADD THÊM QT",
            "cnqt":      "CHẤP NHẬN QT",
            "taodata":   "TẠO DATA",
        }
        parts = " → ".join(_names.get(t, t.upper()) for t in selected)
        self.start_btn.config(text=f"▶  {parts}")

    # ── Start / Stop ───────────────────────────────────────
    def _start(self):
        # Kiểm tra tác vụ đã chọn
        selected = [tid for tid, v in self.task_vars.items() if v.get()]
        if not selected:
            messagebox.showwarning("Chưa chọn tác vụ",
                                   "Hãy tick ít nhất 1 tác vụ ở panel bên trái.")
            return
        _READY = {"login", "create2fa", "change2fa", "channel", "getlink", "banner",
                  "rmadmin", "addqtth", "addqtonly", "cnqt", "taodata"}
        _not_ready = [t for t in selected if t not in _READY]
        if _not_ready:
            _names = {"changepw": "Đổi mật khẩu"}
            names = ", ".join(f"«{_names.get(t, t)}»" for t in _not_ready)
            messagebox.showinfo("Đang phát triển",
                f"Tính năng {names} chưa sẵn sàng — sẽ có ở phiên bản tiếp theo!\n"
                "Bỏ tick các tác vụ đó để tiếp tục.")
            return

        gpm_ok = False
        try:
            gpm_ok = GPMClient().is_running()
        except Exception:
            gpm_ok = False
        if not gpm_ok:
            messagebox.showwarning("GPM chưa kết nối",
                                   "Hãy mở GPM-Login rồi chạy lại.")
            return
        if not self.profiles and not self.create_new_var.get():
            messagebox.showwarning(
                "Thiếu profile GPM",
                "GPM chưa có profile nào. Bật 'Tạo profile mới nếu không có' "
                "hoặc tạo sẵn profile trong GPM.")
            return
        groups, err = self._build_run_plan()
        if err:
            messagebox.showwarning("Chưa chạy được", err)
            self._log(f"✗ {err}", "err")
            return

        self.stop_flag = False
        self.start_btn.config(state=tk.DISABLED)
        self.main_nb.select(1)   # tự chuyển sang tab Kết quả & Log
        with self._results_lock:
            self.results = []          # reset kết quả cho lần chạy mới
        total_acc = sum(len(v) for v in groups.values())
        self._progress_total = total_acc
        self._progress_done = 0
        self.root.after(0, lambda t=total_acc: [
            self.progress_lbl.config(text=f"0 / {t} tài khoản"),
            self.progress_lbl2.config(text=f"0 / {t} tài khoản")])
        self._log(f"\n{'═'*60}", "muted")
        self._log(f"Bắt đầu: {len(groups)} key / {total_acc} tài khoản.", "ok")

        # Lưu cấu hình Google Sheet (URL/secret) để lần sau tự nạp
        try:
            self._gsheet_save_cfg()
            if self._gsheet_on():
                _m = self._machine_name()
                if _m:
                    self._log(f"🔗 Google Sheet: BẬT — máy '{_m}' → đọc tab NHAP_{_m}, "
                              f"ghi tab KETQUA_{_m}.", "ok")
                else:
                    self._log("🔗 Google Sheet: BẬT — (chưa đặt Tên máy → dùng tab mặc định "
                              "INPUT/KETQUA). Đặt Tên máy nếu chạy nhiều máy chung 1 sheet.", "ok")
        except Exception:
            pass

        # ── CHẾ ĐỘ TEST: lưu list account + nạp/khởi tạo tiến độ ──
        if self._test_on():
            try:
                self._session_save_accounts(self._acc_get_text())
                # Lưu tác vụ đã chọn để restart resume tự khôi phục
                try:
                    import json as _jt
                    _tp = Path(__file__).parent / "data" / "_test_tasks.json"
                    _tp.parent.mkdir(exist_ok=True)
                    _tp.write_text(_jt.dumps(self._selected_ordered()),
                                   encoding="utf-8")
                except Exception:
                    pass
                self._progress_load()
                _h = self._accounts_hash(groups)
                if self._progress.get("list_hash") != _h:
                    self._progress = {"done": set(), "skip": set(), "list_hash": _h}
                    self._progress_save()
                    self._log("🧪 TEST: danh sách MỚI → reset tiến độ.", "warn")
                else:
                    _nd = len(self._progress.get("done", set()))
                    _ns = len(self._progress.get("skip", set()))
                    self._log(f"🧪 TEST: chạy tiếp — đã xong {_nd}, bỏ qua {_ns} (sẽ skip).", "warn")
            except Exception as _e:
                self._log(f"🧪 TEST init lỗi: {_e}", "err")

        region = self.region_var.get()
        self._tree_clear()   # xoá bảng kết quả cũ

        def _finish():
            self.root.after(0, lambda: self.start_btn.config(state=tk.NORMAL))
            self.root.after(0, lambda: [
                self.progress_lbl.config(text=f"✓ {self._progress_done} / {self._progress_total} tài khoản"),
                self.progress_lbl2.config(text=f"✓ {self._progress_done} / {self._progress_total} tài khoản")])
            self._set_status("Hoàn thành")
            self._log(f"\n{'═'*60}", "muted")
            self._log("✔ Hoàn thành tất cả.", "ok")
            self._save_results(open_after=False)   # tự lưu CSV kết quả
            # CHẾ ĐỘ TEST: nếu chạy hết mà KHÔNG bị dừng giữa chừng → test xong, xoá phiên
            if self._test_on() and not self.stop_flag:
                try:
                    self._progress_path().unlink(missing_ok=True)
                    self._session_path().unlink(missing_ok=True)
                    self._log("🧪 TEST: đã chạy hết list — xoá phiên test (lần mở sau về bình thường).", "ok")
                except Exception:
                    pass

        # ── ĐỒNG BỘ GIỜ TỪ SERVER (quan trọng cho TOTP) ───────
        # Nếu đồng hồ máy sai giờ, mã 2FA sinh ra sẽ lệch → Google từ chối. Lấy giờ thật từ
        # server Google 1 lần trước khi chạy, bù cho mọi mã TOTP.
        try:
            _off = sync_time_offset(self._log)
            if abs(_off) > 60:
                self._log(f"⚠ ĐỒNG HỒ MÁY LỆCH {_off:+.0f} GIÂY ({_off/3600:+.1f} giờ)! "
                          f"Tool đã tự bù giờ cho 2FA. Nên chỉnh lại giờ Windows (tự đồng bộ).", "err")
        except Exception as _te:
            self._log(f"⚠ Không đồng bộ được giờ: {_te}", "muted")

        # ── Queue + dynamic thread pool ────────────────────────
        # HÀNG ĐỢI ACCOUNT CHUNG — mọi luồng cùng rút account kế tiếp → chạy đều tới
        # account CUỐI CÙNG, không còn cảnh cuối lượt chỉ 1 luồng chạy.
        self._acc_queue = _queue.Queue()
        for _k, _accs in groups.items():
            for _a in _accs:
                self._acc_queue.put(_a)
        # _task_queue: mỗi KEY = 1 worker (mỗi worker gắn 1 key, cùng rút từ _acc_queue).
        self._task_queue = _queue.Queue()
        for idx, key in enumerate(groups.keys(), 1):
            self._task_queue.put((idx, key))

        self._active_workers = 0
        _w_lock = threading.Lock()   # local lock cho lần chạy này

        def worker():
            with _w_lock:
                self._active_workers += 1
            _dec = False
            try:
                while not self.stop_flag:
                    # Tự thoát nếu người dùng giảm số luồng
                    try:
                        _tgt = max(1, min(20, int(self.concurrency_var.get())))
                    except Exception:
                        _tgt = 1
                    with _w_lock:
                        if self._active_workers > _tgt:
                            self._active_workers -= 1
                            _dec = True
                            return
                    # Lấy 1 KEY (mỗi worker gắn 1 key, rồi rút account từ hàng đợi chung)
                    try:
                        _idx, _key = self._task_queue.get_nowait()
                    except _queue.Empty:
                        break
                    try:
                        self._run_key_thread(_key, region, _idx)
                    finally:
                        self._task_queue.task_done()
            finally:
                if not _dec:
                    with _w_lock:
                        self._active_workers -= 1

        def spawner():
            """Spawn thêm worker khi người dùng tăng 'Số luồng' trong khi chạy."""
            while not self.stop_flag and not self._task_queue.empty():
                try:
                    _tgt = max(1, min(20, int(self.concurrency_var.get())))
                except Exception:
                    _tgt = 1
                with _w_lock:
                    _need = max(0, _tgt - self._active_workers)
                for _ in range(_need):
                    if not self._task_queue.empty() and not self.stop_flag:
                        threading.Thread(target=worker, daemon=True).start()
                time.sleep(0.5)

        def _watch():
            self._task_queue.join()   # chờ task_done() được gọi cho mọi item
            _finish()

        # Spawn workers ban đầu
        try:
            n_threads = max(1, min(20, int(self.concurrency_var.get())))
        except Exception:
            n_threads = 1
        n_threads = min(n_threads, self._task_queue.qsize())
        self._log(f"▶ Chạy {n_threads} luồng song song "
                  f"(tăng/giảm 'Số luồng' ngay trong khi chạy để thay đổi).", "muted")
        for _ in range(n_threads):
            threading.Thread(target=worker, daemon=True).start()
        threading.Thread(target=spawner, daemon=True).start()
        threading.Thread(target=_watch, daemon=True).start()

    def _stop(self):
        self.stop_flag = True
        self._log("⬛ Đã yêu cầu dừng... (chờ luồng hiện tại kết thúc bước đang chạy)", "warn")
        # Drain queue để _watch (q.join) không bị block mãi
        if self._task_queue is not None:
            import queue as _q
            while True:
                try:
                    self._task_queue.get_nowait()
                    self._task_queue.task_done()
                except _q.Empty:
                    break

    # ── 1 luồng = 1 key ────────────────────────────────────
    def _run_key_thread(self, key: str, region: str, tnum: int):
        try:
            asyncio.run(self._process_key(key, region, tnum))
        except Exception as e:
            self._log(f"[L{tnum}] ✗ Luồng lỗi: {e}", "err")

    def _wait_rotation(self, next_at_ms, tag: str):
        """Chờ tới mốc nextRequestAt (ms) mới được đổi proxy. Kiểm tra stop_flag."""
        if not next_at_ms:
            return
        target = next_at_ms / 1000.0
        while not self.stop_flag:
            remain = target - time.time()
            if remain <= 0:
                return
            wait = min(remain, 5)
            self._set_status(f"{tag} chờ đổi IP {int(remain)}s")
            time.sleep(wait)

    @staticmethod
    def _parse_wait_seconds(msg: str) -> int:
        """Trích số giây cần chờ từ thông báo 'Proxy chưa đến hạn đổi. Đợi lại sau 46 giây'."""
        import re
        m = re.search(r'(\d+)\s*gi[aâ]y', msg) or re.search(r'(\d+)\s*s', msg) or re.search(r'(\d+)', msg)
        return int(m.group(1)) if m else 0

    def _sleep_stop(self, secs: float, tag: str = ""):
        end = time.time() + secs
        while time.time() < end and not self.stop_flag:
            if tag:
                self._set_status(f"{tag} chờ đổi IP {int(end - time.time())}s")
            time.sleep(min(2, max(0, end - time.time())))

    def _acquire_proxy(self, key: str, region: str, first: bool, tag: str):
        """Lấy proxy cho account. Luôn thử xoay IP mới (/new) trước để tối đa đa dạng IP.
        Chỉ fallback /current khi đang cooldown VÀ là account đầu của key (không chờ)."""
        for attempt in range(8):
            if self.stop_flag:
                return None
            try:
                p = self.kiot.get_new(key, region)
                return p
            except KiotError as e:
                emsg = str(e)
                wait = self._parse_wait_seconds(emsg)
                if "chưa đến hạn" in emsg.lower() or wait:
                    if first:
                        # Cooldown chưa hết + account đầu → dùng IP hiện tại, không chờ
                        try:
                            p = self.kiot.get_current(key)
                            if p and p.get("host"):
                                self._log(f"{tag} Cooldown chưa hết → dùng IP hiện tại", "muted")
                                return p
                        except Exception:
                            pass
                    # Account tiếp theo (hoặc get_current lỗi) → chờ hết cooldown rồi xoay
                    w = min(max(wait, 5), 90) + 2
                    self._log(f"{tag} ⏳ Chờ {w}s để đổi IP… ({emsg})", "warn")
                    self._sleep_stop(w, tag)
                else:
                    self._log(f"{tag} ⚠ Lấy IP lỗi: {emsg}", "warn")
                    self._sleep_stop(5, tag)
            except Exception as e:
                self._log(f"{tag} ✗ Kết nối kiotproxy lỗi: {e}", "err")
                return None
        return None

    # ══════════ GOOGLE SHEET (đọc account + ghi kết quả online) ══════════
    def _gsheet_cfg_path(self):
        return Path(__file__).parent / "data" / "_gsheet.json"

    def _gsheet_load_cfg(self) -> dict:
        import json
        try:
            p = self._gsheet_cfg_path()
            if p.exists():
                return json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            pass
        return {}

    def _gsheet_save_cfg(self):
        import json
        try:
            p = self._gsheet_cfg_path()
            p.parent.mkdir(exist_ok=True)
            p.write_text(json.dumps({
                "url": self.gsheet_url_var.get().strip(),
                "secret": self.gsheet_secret_var.get().strip(),
                "machine": self._machine_name(),
                "enabled": bool(self.gsheet_on_var.get()),
                "nolocal": bool(self.gsheet_nolocal_var.get()),
                "test": bool(getattr(self, "test_mode_var", None) and self.test_mode_var.get()),
                "threads": int(getattr(self, "concurrency_var", None).get()
                               if getattr(self, "concurrency_var", None) else 1),
            }, ensure_ascii=False, indent=1), encoding="utf-8")
        except Exception:
            pass

    def _gsheet_on(self) -> bool:
        try:
            return bool(self.gsheet_on_var.get()) and bool(self.gsheet_url_var.get().strip())
        except Exception:
            return False

    @staticmethod
    def _sanitize_machine(name: str) -> str:
        """Làm sạch tên máy để hợp lệ làm TÊN TAB Google Sheet.
        Bỏ ký tự tab-name cấm ( [ ] * ? / \\ : ), bỏ dấu cách 2 đầu, thay khoảng trắng
        trong tên bằng '-'. Giữ chữ/số/gạch/underscore + Unicode (Sheet cho phép)."""
        import re as _re
        s = (name or "").strip()
        s = _re.sub(r"[\[\]\*\?/\\:]", "", s)     # ký tự cấm trong tên tab
        s = _re.sub(r"\s+", "-", s)               # khoảng trắng -> gạch
        return s[:80]                              # tên tab tối đa 100, để dư

    def _machine_name(self) -> str:
        try:
            return self._sanitize_machine(self.machine_var.get())
        except Exception:
            return ""

    def _input_tab(self) -> str:
        """Tên tab INPUT của máy này (rỗng nếu chưa đặt tên máy → dùng tab mặc định)."""
        m = self._machine_name()
        return f"NHAP_{m}" if m else ""

    def _gsheet_load_accounts(self):
        if gsheet_client is None:
            messagebox.showwarning("Thiếu module", "Không import được gsheet_client.py")
            return
        # CHỐNG BẤM CHỒNG: nếu đang tải thì bỏ qua (tránh spawn nhiều luồng request → lag).
        if getattr(self, "_gsheet_loading", False):
            return
        self._gsheet_loading = True
        try:
            self._gsheet_load_btn.config(text="⏳ Đang nạp…", state=tk.DISABLED)
        except Exception:
            pass
        self._gsheet_save_cfg()
        url = self.gsheet_url_var.get().strip()
        sec = self.gsheet_secret_var.get().strip()
        _tab = self._input_tab()

        def _reset_btn():
            self._gsheet_loading = False
            try:
                self._gsheet_load_btn.config(text="📥 Nạp account từ Sheet", state=tk.NORMAL)
            except Exception:
                pass

        def _do():
            try:
                accs = gsheet_client.fetch_accounts(url, sec, _tab)

                def _fill():
                    self.acc_text.delete("1.0", tk.END)
                    self.acc_text.insert("1.0", "\n".join(accs) + ("\n" if accs else ""))
                    self._acc_ph_on = False   # QUAN TRỌNG: tắt cờ placeholder để _acc_get_text đọc thật
                    self.acc_text.config(fg=self.COLOR_TEXT)
                    self._update_line_count()
                    self._log(f"📥 Nạp {len(accs)} account từ Google Sheet.", "ok")
                    _reset_btn()
                self.root.after(0, _fill)
            except Exception as e:
                _emsg = str(e)

                def _err():
                    self._log(f"✗ Nạp account từ Sheet lỗi: {_emsg}", "err")
                    _reset_btn()
                self.root.after(0, _err)
        threading.Thread(target=_do, daemon=True).start()

    def _gsheet_test(self):
        if gsheet_client is None:
            messagebox.showwarning("Thiếu module", "Không import được gsheet_client.py")
            return
        self._gsheet_save_cfg()
        url = self.gsheet_url_var.get().strip()
        sec = self.gsheet_secret_var.get().strip()

        _tab = self._input_tab()

        def _do():
            ok, msg = gsheet_client.test_connection(url, sec, _tab)
            self.root.after(0, lambda: self._log(
                ("✓ " if ok else "✗ ") + "Google Sheet: " + msg, "ok" if ok else "err"))
        threading.Thread(target=_do, daemon=True).start()

    def _gsheet_push(self, row: dict):
        """Ghi 1 dòng kết quả lên Sheet (UPSERT theo email). Gọi trong worker thread."""
        if not self._gsheet_on() or gsheet_client is None:
            return
        try:
            gsheet_client.write_result(self.gsheet_url_var.get().strip(),
                                       self.gsheet_secret_var.get().strip(), row,
                                       self._input_tab())
        except Exception as e:
            self._log(f"⚠ Ghi Sheet lỗi ({row.get('email','')}): {str(e)[:70]}", "warn")

    def _selected_ordered(self) -> list:
        """Tác vụ ĐƯỢC CHỌN theo ĐÚNG THỨ TỰ người dùng click.
        (Tác vụ chọn mà chưa có trong _task_order → thêm cuối theo thứ tự mặc định.)"""
        out = [t for t in self._task_order
               if t in self.task_vars and self.task_vars[t].get()]
        for t, v in self.task_vars.items():
            if v.get() and t not in out:
                out.append(t)
        if "login" in out:            # 'Đăng nhập' luôn ở ĐẦU (là bước nền)
            out.remove("login")
            out.insert(0, "login")
        return out

    # ══════════ CHẾ ĐỘ TEST: tiến độ + phiên list account ══════════
    def _test_on(self) -> bool:
        try:
            return bool(self.test_mode_var.get())
        except Exception:
            return False

    def _progress_path(self):
        return Path(__file__).parent / "data" / "_test_progress.json"

    def _session_path(self):
        return Path(__file__).parent / "data" / "_test_accounts.txt"

    def _progress_load(self):
        import json
        self._progress = {"done": set(), "skip": set(), "list_hash": ""}
        try:
            p = self._progress_path()
            if p.exists():
                d = json.loads(p.read_text(encoding="utf-8"))
                self._progress["done"] = set(d.get("done", []))
                self._progress["skip"] = set(d.get("skip", []))
                self._progress["list_hash"] = d.get("list_hash", "")
        except Exception:
            pass

    def _progress_save(self):
        import json
        # KHÓA: tránh 2 luồng vừa .add() vừa sorted()/ghi file cùng lúc (mất tiến độ / hỏng file).
        with self._progress_lock:
            try:
                p = self._progress_path()
                p.parent.mkdir(exist_ok=True)
                p.write_text(json.dumps({
                    "done": sorted(self._progress.get("done", set())),
                    "skip": sorted(self._progress.get("skip", set())),
                    "list_hash": self._progress.get("list_hash", ""),
                }, ensure_ascii=False, indent=1), encoding="utf-8")
            except Exception:
                pass

    def _progress_add(self, kind: str, email: str):
        with self._progress_lock:   # đồng bộ đa luồng
            try:
                if not hasattr(self, "_progress"):
                    self._progress_load()
                self._progress.setdefault(kind, set()).add(email)
                self._progress_save()
            except Exception:
                pass

    def _test_should_skip(self, email: str) -> bool:
        if not self._test_on():
            return False
        pr = getattr(self, "_progress", None) or {}
        return email in pr.get("done", set()) or email in pr.get("skip", set())

    def _accounts_hash(self, groups) -> str:
        import hashlib
        emails = []
        for _k, accs in groups.items():
            emails += [a.get("email", "") for a in accs]
        return hashlib.md5("|".join(sorted(emails)).encode("utf-8")).hexdigest()[:12]

    def _session_save_accounts(self, text: str):
        try:
            p = self._session_path()
            p.parent.mkdir(exist_ok=True)
            p.write_text(text, encoding="utf-8")
        except Exception:
            pass

    def _session_load_accounts(self):
        """Nạp lại list account đã lưu (để restart không mất danh sách)."""
        try:
            p = self._session_path()
            if p.exists():
                txt = p.read_text(encoding="utf-8")
                if txt.strip():
                    self.acc_text.delete("1.0", tk.END)
                    self.acc_text.insert("1.0", txt)
                    self._acc_ph_on = False   # tắt cờ placeholder để _acc_get_text đọc thật
                    self.acc_text.config(fg=self.COLOR_TEXT)
                    # Khôi phục tác vụ đã chọn (để restart resume không phải tick lại tay)
                    try:
                        import json as _json
                        tp = Path(__file__).parent / "data" / "_test_tasks.json"
                        if tp.exists():
                            _tlist = _json.loads(tp.read_text(encoding="utf-8"))
                            _tasks = set(_tlist)
                            # DỰNG LẠI thứ tự click đã lưu (danh sách có thứ tự trong file).
                            self._task_order = [t for t in _tlist if t in self.task_vars]
                            if _tasks:
                                for _tid, _v in self.task_vars.items():
                                    _v.set(_tid in _tasks)
                                    # VẼ LẠI checkbox cho khớp biến (tránh lệch: ô bỏ tick
                                    # nhưng vẫn chạy vì biến = True sau khôi phục phiên test).
                                    try:
                                        if getattr(self, "_apply_row_state_fn", None):
                                            self._apply_row_state_fn(_tid, _tid in _tasks)
                                    except Exception:
                                        pass
                                self._refresh_start_btn()
                    except Exception:
                        pass
                    return True
        except Exception:
            pass
        return False

    # Trạng thái = lỗi DỮ LIỆU (không sửa bằng code → tự đánh dấu bỏ qua)
    _DATA_ERR_MARKERS = ("SAI_MK", "SAI MẬT KHẨU", "MẬT KHẨU SAI", "DISABLED",
                         "VÔ HIỆU", "CAPTCHA", "PROXY", "SAI_EMAIL", "SAI EMAIL",
                         "KÊNH DIE", "QUÁ NHIỀU", "KHÔNG LẤY ĐƯỢC IP")

    def _test_after_record(self, email: str, status: str, detail: str, channel: str):
        """Chế độ TEST: đánh dấu done/skip + DỪNG khi gặp lỗi."""
        if not self._test_on():
            return
        # Đang DỪNG giữa chừng → KHÔNG đánh dấu done/skip (account chưa chắc xong hết) →
        # lần chạy lại (resume) sẽ LÀM LẠI account này thay vì bỏ qua.
        if self.stop_flag:
            return
        su = (status or "").upper()
        det = (detail or "")
        ch = (channel or "")
        # Thành công hoàn toàn: THÀNH CÔNG, không có 'THẤT BẠI/LỖI', kênh không lỗi tạo
        _full = (su.strip() == "THÀNH CÔNG"
                 and "THẤT BẠI" not in su and "LỖI" not in su
                 and "không tạo được" not in det.lower()
                 and "không đọc được" not in det.lower()
                 and "KÊNH DIE" not in ch.upper())
        if _full:
            self._progress_add("done", email)
            return
        # YÊU CẦU NGƯỜI DÙNG: CHẠY HẾT CẢ LƯỢT — KHÔNG DỪNG vì bất kỳ lỗi nào.
        # Account nào không THÀNH CÔNG (lỗi login/captcha/verify/đổi-tạo 2FA/kênh…) đều
        # được đánh dấu 'skip' + ghi log lỗi để CUỐI CÙNG người dùng vào check. Không set
        # stop_flag → tool tự chạy tới account cuối rồi mới báo "Hoàn thành tất cả".
        self._progress_add("skip", email)
        self._log(f"🧪 {email} LỖI ({(status or '')[:45]}) → ghi nhận, chạy tiếp (không dừng).", "warn")

    async def _process_key(self, key: str, region: str, tnum: int):
        tag = f"[L{tnum} {key[:6]}…]"
        gpm = GPMClient()
        last_next_at = None
        self._log(f"{tag} Bắt đầu — rút account từ hàng đợi chung.", "muted")

        i = 0
        while not self.stop_flag:
            # HÀNG ĐỢI CHUNG: rút account KẾ TIẾP. Hết account thì luồng này kết thúc,
            # nhưng các luồng khác vẫn tiếp tục cho tới account cuối → không cụt luồng.
            try:
                acc = self._acc_queue.get_nowait()
            except _queue.Empty:
                break
            i += 1

            email = acc["email"]
            # CHẾ ĐỘ TEST: bỏ qua account đã xong/đã đánh dấu ở lần chạy trước
            if self._test_should_skip(email):
                self._log(f"{tag} ⏭ Bỏ qua {email} (đã xong/đã đánh dấu ở lần trước)", "muted")
                continue
            # Lưu creds để _record xuất ra file đầy đủ (mail|pass|recovery|2fa)
            self._acc_creds[email] = acc
            self._log(f"\n{tag} (#{i}) ── {email} ──", "muted")
            self._set_status(f"{tag} {email}")
            self._tree_upsert(email, luong=f"L{tnum}", status="⏳ Đang lấy proxy…",
                              tag="run", matkhau=acc.get("password", ""),
                              recovery=acc.get("recovery", ""), twofa=acc.get("totp", ""))

            # 1) Lấy IP: account đầu của key -> dùng IP hiện tại (/current);
            #    account sau -> xoay IP mới (/new, chờ hết cooldown).
            proxy = self._acquire_proxy(key, region, first=(i == 1), tag=tag)
            if not proxy or not proxy.get("host"):
                self._log(f"{tag} ✗ Không lấy được proxy — bỏ qua {email}", "err")
                self._record(key, email, "LỖI PROXY (không lấy được IP)")
                continue

            last_next_at = proxy.get("next_at")
            host = proxy["host"]
            port = proxy["http_port"]
            location = proxy.get("location", "")
            ip_str = f"{host}:{port}"      # IP proxy gán cho kênh này
            self._log(f"{tag} 🌐 IP: {host}:{port} @ {location} "
                      f"(ttc={proxy['ttc']}s)", "ok")
            self._tree_upsert(email, ip=f"{host}:{port} @ {location}",
                              status="⏳ Đang đăng nhập…", tag="run")

            # 2) Test proxy + lấy IP thật khi ra internet -------------------
            exit_ip = host   # mặc định = host proxy
            if self.test_proxy_var.get():
                ok, msg = gpm.test_proxy(host, port, protocol="http")
                if ok:
                    self._log(f"{tag} ✓ {msg}", "ok")
                    import re as _re2
                    m = _re2.search(r"(\d{1,3}(?:\.\d{1,3}){3})", msg or "")
                    if m:
                        exit_ip = m.group(1)
                else:
                    self._log(f"{tag} ⚠ Proxy có thể lỗi: {msg} (vẫn tiếp tục)", "warn")
            # Chuỗi IP ghi vào kết quả (chỉ để HIỂN THỊ): IP thật (cổng) @ vị trí
            ip_str = f"{exit_ip}:{port}" + (f" @ {location}" if location else "")
            # Chuỗi proxy SẠCH truyền cho GPM (KHÔNG kèm ' @ vị trí', nếu không
            # GPM parse port = 0 -> proxy hỏng -> profile mở lỗi).
            raw_proxy_clean = f"{host}:{port}"

            # 3) Tìm / tạo profile -----------------------------------------
            created_with_proxy = False
            profile_id, profile_name = self._find_profile(email)
            if not profile_id:
                if self.create_new_var.get():
                    try:
                        name = email.split("@")[0]
                        with GPM_UI_LOCK:   # tránh nhiều luồng cùng thao tác GPM
                            profile_id = gpm.create_profile(name=name, raw_proxy=raw_proxy_clean)
                        if not profile_id:
                            raise RuntimeError("create_profile trả về None")
                        profile_name = name
                        created_with_proxy = True   # đã set proxy khi tạo (API v3)
                        self._log(f"{tag} ✓ Tạo profile mới: '{name}' (kèm proxy)", "ok")
                        self.root.after(0, self._load_profiles)
                    except Exception as e:
                        self._log(f"{tag} ✗ Không tạo được profile: {e}", "err")
                        self._record(key, email, "LỖI TẠO PROFILE", ip_str, str(e))
                        continue
                else:
                    self._log(f"{tag} ✗ Không tìm thấy profile cho '{email}' "
                              f"(bật 'Tạo profile mới' hoặc dùng chế độ chọn thủ công)", "err")
                    self._record(key, email, "THIẾU PROFILE GPM", ip_str)
                    continue

            # 4) Ghi proxy vào GPM (nếu chưa set khi tạo) ------------------
            if created_with_proxy:
                self._log(f"{tag} ✓ Proxy đã set khi tạo profile", "ok")
            else:
                try:
                    # đóng profile nếu đang mở để đảm bảo nạp proxy mới
                    gpm_stop(profile_id)
                    time.sleep(1)
                    with DB_LOCK:
                        ok, msg = gpm.update_proxy(profile_id, host, port, protocol="http")
                    if not ok:
                        self._log(f"{tag} ✗ Ghi proxy vào GPM lỗi: {msg}", "err")
                        self._record(key, email, "LỖI SET PROXY", ip_str, msg)
                        continue
                    self._log(f"{tag} ✓ Đã set proxy vào profile '{profile_name}'", "ok")
                except Exception as e:
                    self._log(f"{tag} ✗ Lỗi set proxy: {e}", "err")
                    self._record(key, email, "LỖI SET PROXY", ip_str, str(e))
                    continue

            # 5) Mở profile (có delay + retry vì profile vừa tạo) ----------
            if created_with_proxy:
                # Chờ GPM lưu xong profile: poll list tới khi thấy id (tối đa ~8s)
                for _w in range(8):
                    if self.stop_flag:
                        break
                    time.sleep(1)
                    try:
                        ids = {p.get("id") or p.get("profileId") for p in (gpm_list() or [])}
                        if profile_id in ids:
                            self._log(f"{tag} ✓ Profile đã lưu trong GPM", "muted")
                            break
                    except Exception:
                        pass
                else:
                    self._log(f"{tag} ⚠ Chưa thấy profile trong GPM sau 8s "
                              f"(vẫn thử mở)", "warn")
            ws_url = None
            last_err = ""
            for attempt in range(6):
                if self.stop_flag:
                    break
                try:
                    self._log(f"{tag} Mở profile... (lần {attempt+1}/6)", "muted")
                    ws_url = gpm_start(profile_id)
                    break
                except Exception as e:
                    last_err = str(e)
                    self._log(f"{tag} ⚠ Mở profile lỗi (có thể GPM đang tải browser): "
                              f"{str(e)[:70]}", "warn")
                    self._sleep_stop(10, tag)
            if not ws_url:
                self._log(f"{tag} ✗ Không mở được profile sau 6 lần", "err")
                self._record(key, email, "LỖI MỞ PROFILE", ip_str, last_err)
                continue
            self._log(f"{tag} ✓ Browser đã mở", "ok")

            # 6) Đăng nhập Google (nếu proxy lỗi mạng -> đổi IP, chạy lại) -
            success = False
            channel = None
            err_str = ""
            for login_try in range(1, 4):   # tối đa 3 lần (đổi IP giữa các lần)
                if self.stop_flag:
                    break
                try:
                    # HẠN GIỜ 5 phút: nếu login treo (proxy rớt/UI đổi) thì bỏ qua, KHÔNG kẹt luồng.
                    success, result, channel = await asyncio.wait_for(
                        do_google_login(
                            ws_url, email, acc["password"], acc["recovery"],
                            totp_secret=acc.get("totp", ""),
                            log_fn=lambda m: self._log(f"{tag}{m}", "muted"),
                        ), timeout=300)
                    err_str = "" if success else str(result)
                except asyncio.TimeoutError:
                    success = False
                    err_str = "Đăng nhập QUÁ 5 PHÚT (treo) — bỏ qua để không kẹt luồng"
                    channel = None
                    self._log(f"{tag} ✗ {err_str}", "err")
                except Exception as e:
                    success = False
                    err_str = str(e)
                    channel = None

                if success or not self._is_proxy_error(err_str):
                    break
                if login_try >= 3:
                    break
                # Proxy lỗi mạng -> đổi IP mới rồi chạy lại account này
                self._log(f"{tag} ⚠ Proxy lỗi mạng → đổi IP, chạy lại (lần {login_try})", "warn")
                self._tree_upsert(email, status="🔁 Đổi IP, chạy lại…", tag="run")
                try:
                    self._sleep_stop(3, tag)
                    np = self.kiot.get_new(key, region)   # lấy IP mới
                    if np and np.get("host"):
                        host = np["host"]; port = np["http_port"]
                        location = np.get("location", "")
                        ip_str = f"{host}:{port} @ {location}"
                        self._log(f"{tag} 🌐 IP mới: {host}:{port} @ {location}", "ok")
                        self._tree_upsert(email, ip=ip_str, status="⏳ Đang đăng nhập…", tag="run")
                        gpm_stop(profile_id)
                        time.sleep(1)
                        with DB_LOCK:
                            gpm.update_proxy(profile_id, host, port, protocol="http")
                        ws_url = gpm_start(profile_id)
                    else:
                        break
                except Exception as e:
                    self._log(f"{tag} ⚠ Đổi IP lỗi: {str(e)[:60]}", "warn")
                    break

            # ── Tác vụ sau login ───────────────────────────────────────────
            new_totp = acc.get("totp", "")   # sẽ cập nhật nếu tạo/đổi 2FA
            task_status = "THÀNH CÔNG"
            task_detail = "Đăng nhập OK"
            _banner_ok = None   # khởi tạo SỚM (tránh NameError nếu bấm Dừng giữa chừng)
            _rmadmin_ok = None
            _addqt_ok = None
            _addqt_note = ""    # vd "2FA 7 ngày" khi Google chặn xác minh
            _addqtonly_ok = None
            _addqtonly_note = ""
            _cnqt_ok = None
            _cnqt_note = ""
            _taodata_ok = None
            _taodata_detail = ""   # tóm tắt bước nào OK / lỗi

            if success and not self.stop_flag:
                _sel = self._selected_ordered()   # tác vụ theo ĐÚNG THỨ TỰ click

                # ══ Mỗi tác vụ = 1 hàm; GỌI theo thứ tự click ở cuối khối ══
                # ── Tạo / Đổi 2FA — tự quyết theo account ─────────────────
                async def _run_2fa():
                    nonlocal task_status, task_detail, new_totp
                    _has_2fa = bool(acc.get("totp", "").strip())
                    if _has_2fa:
                        # Đã có 2FA → đổi
                        self._log(f"{tag} 🔐 Account đã có 2FA → Đổi 2FA…", "muted")
                        self._tree_upsert(email, status="⏳ Đổi 2FA…", tag="run")
                        try:
                            ok2, secret, msg = await asyncio.wait_for(do_change_2fa(
                                ws_url, email, acc["password"],
                                old_totp_secret=acc.get("totp", ""),
                                recovery_email=acc.get("recovery", ""),
                                log_fn=lambda m: self._log(f"{tag}{m}", "muted")),
                                timeout=300)
                            if ok2:
                                old_totp = acc.get("totp", "")   # lưu 2FA cũ
                                new_totp = secret
                                self._acc_creds.setdefault(email, acc)["old_totp"] = old_totp
                                self._log(f"{tag} ✅ Đổi 2FA OK — secret mới: {secret}", "ok")
                                task_detail = "Đăng nhập OK | Đổi 2FA OK"
                            else:
                                self._log(f"{tag} ⚠ Đổi 2FA thất bại: {msg}", "warn")
                                if str(msg).startswith("DEVICE_CODE"):
                                    task_status = "ĐĂNG NHẬP OK / DÍNH THIẾT BỊ CŨ (duyệt tay)"
                                else:
                                    task_status = "ĐĂNG NHẬP OK / ĐỔI 2FA THẤT BẠI"
                                task_detail = msg
                        except _DeviceCodeChallenge:
                            self._log(f"{tag} ⛔ DÍNH THIẾT BỊ CŨ — Google đòi Mã bảo mật từ thiết bị.", "warn")
                            task_status = "ĐĂNG NHẬP OK / DÍNH THIẾT BỊ CŨ (duyệt tay)"
                            task_detail = ("DEVICE_CODE | Google đòi 'Mã bảo mật' từ thiết bị đã "
                                           "đăng ký — không tự động được, duyệt tay bằng thiết bị đó.")
                        except asyncio.TimeoutError:
                            self._log(f"{tag} ✗ Đổi 2FA QUÁ 5 PHÚT (treo) — bỏ qua.", "err")
                            task_status = "ĐĂNG NHẬP OK / ĐỔI 2FA QUÁ GIỜ"
                            task_detail = "Đổi 2FA quá 5 phút (treo)"
                        except Exception as e2:
                            self._log(f"{tag} ✗ Lỗi đổi 2FA: {e2}", "err")
                            task_status = "ĐĂNG NHẬP OK / LỖI ĐỔI 2FA"
                            task_detail = str(e2)[:120]
                    else:
                        # Chưa có 2FA → tạo mới
                        self._log(f"{tag} 🔑 Account chưa có 2FA → Tạo 2FA…", "muted")
                        self._tree_upsert(email, status="⏳ Tạo 2FA…", tag="run")
                        try:
                            ok2, secret, msg = await asyncio.wait_for(do_create_2fa(
                                ws_url, email, acc["password"],
                                recovery_email=acc.get("recovery", ""),
                                log_fn=lambda m: self._log(f"{tag}{m}", "muted")),
                                timeout=300)
                            if ok2:
                                new_totp = secret
                                self._log(f"{tag} ✅ Tạo 2FA OK — secret: {secret}", "ok")
                                task_detail = "Đăng nhập OK | Tạo 2FA OK"
                            else:
                                self._log(f"{tag} ⚠ Tạo 2FA thất bại: {msg}", "warn")
                                if str(msg).startswith("DEVICE_CODE"):
                                    task_status = "ĐĂNG NHẬP OK / DÍNH THIẾT BỊ CŨ (duyệt tay)"
                                else:
                                    task_status = "ĐĂNG NHẬP OK / TẠO 2FA THẤT BẠI"
                                task_detail = msg
                        except _DeviceCodeChallenge:
                            self._log(f"{tag} ⛔ DÍNH THIẾT BỊ CŨ — Google đòi Mã bảo mật từ thiết bị.", "warn")
                            task_status = "ĐĂNG NHẬP OK / DÍNH THIẾT BỊ CŨ (duyệt tay)"
                            task_detail = ("DEVICE_CODE | Google đòi 'Mã bảo mật' từ thiết bị đã "
                                           "đăng ký — không tự động được, duyệt tay bằng thiết bị đó.")
                        except asyncio.TimeoutError:
                            self._log(f"{tag} ✗ Tạo 2FA QUÁ 5 PHÚT (treo) — bỏ qua.", "err")
                            task_status = "ĐĂNG NHẬP OK / TẠO 2FA QUÁ GIỜ"
                            task_detail = "Tạo 2FA quá 5 phút (treo)"
                        except Exception as e2:
                            self._log(f"{tag} ✗ Lỗi tạo 2FA: {e2}", "err")
                            task_status = "ĐĂNG NHẬP OK / LỖI TẠO 2FA"
                            task_detail = str(e2)[:120]

                # ── Lấy link / Tạo kênh YouTube ───────────────────────────
                #  • Chỉ 'getlink'      → chỉ check link (chưa có → báo, KHÔNG tạo)
                #  • Có 'channel'       → chưa có kênh thì TẠO mới rồi lấy link
                #  • Cả 2               → check link → chưa có thì tạo → die thì báo die
                #  (die luôn được báo dù có tạo hay không)
                async def _run_channel():
                    nonlocal channel, task_status, task_detail
                    _create_ch = bool(self.task_vars["channel"].get())   # chỉ tạo khi tick 'Tạo kênh'
                    _lbl = "Lấy/Tạo kênh" if _create_ch else "Lấy link kênh"
                    self._log(f"{tag} 📺 {_lbl} YouTube…", "muted")
                    self._tree_upsert(email, status="⏳ Kênh YouTube…", tag="run")
                    try:
                        okc, ch_url, msgc = await asyncio.wait_for(do_channel(
                            ws_url, email, create_if_missing=_create_ch,
                            log_fn=lambda m: self._log(f"{tag}{m}", "muted")),
                            timeout=240)
                        if okc and ch_url:
                            channel = ch_url
                            self._log(f"{tag} ✅ Kênh: {ch_url} ({msgc})", "ok")
                            task_detail = (task_detail + " | Kênh OK").strip(" |")
                        elif str(msgc).startswith("KÊNH DIE"):
                            channel = "KÊNH DIE"
                            self._log(f"{tag} ☠ KÊNH DIE: {email}", "err")
                            task_status = "ĐĂNG NHẬP OK / KÊNH DIE"
                            task_detail = (task_detail + " | KÊNH DIE").strip(" |")
                        elif "chưa có kênh" in str(msgc).lower():
                            channel = "CHƯA CÓ KÊNH"
                            self._log(f"{tag} ⚠ Chưa có kênh: {email}", "warn")
                            task_detail = (task_detail + " | Chưa có kênh").strip(" |")
                        else:
                            self._log(f"{tag} ⚠ Kênh: {msgc}", "warn")
                            task_detail = (task_detail + f" | Kênh: {msgc}").strip(" |")
                    except asyncio.TimeoutError:
                        self._log(f"{tag} ✗ Kênh QUÁ 4 PHÚT (treo) — bỏ qua.", "err")
                        task_detail = (task_detail + " | Kênh: quá giờ (treo)").strip(" |")
                    except Exception as ec:
                        self._log(f"{tag} ✗ Lỗi kênh: {ec}", "err")
                        task_detail = (task_detail + " | Lỗi kênh").strip(" |")

                # ── Thay ẢNH BÌA kênh YouTube ──────────────────────────────
                async def _run_banner():
                    nonlocal _banner_ok, task_detail
                    _bdir = str(Path(__file__).parent / "anh_bia")
                    self._log(f"{tag} 🖼️ Thay ảnh bìa kênh…", "muted")
                    self._tree_upsert(email, status="⏳ Thay ảnh bìa…", tag="run")
                    try:
                        # HẠN GIỜ CỨNG 5 phút: dù UI YouTube đổi/treo cũng KHÔNG kẹt luồng.
                        okb, _, msgb = await asyncio.wait_for(
                            do_change_banner(
                                ws_url, email, banner_dir=_bdir,
                                log_fn=lambda m: self._log(f"{tag}{m}", "muted")),
                            timeout=300)
                        _banner_ok = bool(okb)
                        if okb:
                            self._log(f"{tag} ✅ {msgb}", "ok")
                            task_detail = (task_detail + " | Ảnh bìa OK").strip(" |")
                        else:
                            self._log(f"{tag} ⚠ Ảnh bìa: {msgb}", "warn")
                            task_detail = (task_detail + f" | Ảnh bìa lỗi: {msgb}").strip(" |")
                    except asyncio.TimeoutError:
                        _banner_ok = False
                        self._log(f"{tag} ✗ Thay ảnh bìa QUÁ 5 PHÚT — bỏ qua để không kẹt luồng.",
                                  "err")
                        task_detail = (task_detail + " | Ảnh bìa: quá giờ (5 phút)").strip(" |")
                    except Exception as eb:
                        _banner_ok = False
                        self._log(f"{tag} ✗ Lỗi thay ảnh bìa: {eb}", "err")
                        task_detail = (task_detail + f" | Lỗi ảnh bìa: {eb}").strip(" |")

                # ── Thoát quản trị kênh (tự gỡ mình khỏi Brand Account) ────
                # AN TOÀN: nếu chạy 'Thay ảnh bìa' TRƯỚC mà bìa THẤT BẠI thì KHÔNG thoát quản trị —
                # tránh mất quyền admin trong khi bìa chưa đổi được (không vào lại kênh để sửa).
                async def _run_rmadmin():
                    nonlocal _rmadmin_ok, task_detail
                    # AN TOÀN (không phụ thuộc thứ tự click): nếu CÓ chọn 'Thay ảnh bìa' mà bìa
                    # CHƯA thành công (đã lỗi HOẶC chưa chạy vì xếp SAU thoát QT) → BỎ QUA thoát QT,
                    # tránh mất quyền admin khi bìa chưa đổi được.
                    _banner_selected = bool(self.task_vars.get("banner") and self.task_vars["banner"].get())
                    if _banner_selected and _banner_ok is not True:
                        if _banner_ok is False:
                            _why = "ảnh bìa THẤT BẠI"
                        else:
                            _why = "ảnh bìa CHƯA chạy (hãy đặt 'Thay ảnh bìa' TRƯỚC 'Thoát quản trị')"
                        self._log(f"{tag} ⛔ BỎ QUA thoát quản trị vì {_why} "
                                  f"(giữ quyền admin để còn sửa bìa).", "warn")
                        task_detail = (task_detail + f" | Bỏ qua thoát QT ({_why})").strip(" |")
                        return
                    self._log(f"{tag} 🚪 Thoát quản trị kênh…", "muted")
                    self._tree_upsert(email, status="⏳ Thoát quản trị…", tag="run")
                    try:
                        okr, _, msgr = await asyncio.wait_for(
                            do_leave_admin(
                                ws_url, email, password=acc.get("password", ""),
                                log_fn=lambda m: self._log(f"{tag}{m}", "muted")),
                            timeout=240)
                        _rmadmin_ok = bool(okr)
                        if okr:
                            self._log(f"{tag} ✅ {msgr}", "ok")
                            task_detail = (task_detail + " | Thoát QT OK").strip(" |")
                        else:
                            self._log(f"{tag} ⚠ Thoát QT: {msgr}", "warn")
                            task_detail = (task_detail + f" | Thoát QT lỗi: {msgr}").strip(" |")
                    except asyncio.TimeoutError:
                        _rmadmin_ok = False
                        self._log(f"{tag} ✗ Thoát quản trị QUÁ 4 PHÚT — bỏ qua để không kẹt luồng.",
                                  "err")
                        task_detail = (task_detail + " | Thoát QT: quá giờ").strip(" |")
                    except Exception as er:
                        _rmadmin_ok = False
                        self._log(f"{tag} ✗ Lỗi thoát quản trị: {er}", "err")
                        task_detail = (task_detail + f" | Lỗi thoát QT: {er}").strip(" |")

                # ── Add QT TH: tạo/kiểm kênh thương hiệu + add owner ───────
                async def _run_addqtth():
                    nonlocal _addqt_ok, _addqt_note, task_detail
                    self._log(f"{tag} ➕ Add QT TH…", "muted")
                    self._tree_upsert(email, status="⏳ Add QT TH…", tag="run")
                    try:
                        oka, _codea, msga = await asyncio.wait_for(
                            do_add_brand_admin(
                                ws_url, email, password=acc.get("password", ""),
                                owner_email=acc.get("owner", ""),   # cột E: mail quản trị
                                # 2FA cho bước xác minh khi move — DÙNG new_totp (secret MỚI nếu 2FA
                                # vừa được tạo trong cùng lần chạy; nếu không thì totp cũ).
                                totp_secret=(new_totp or acc.get("totp", "")),
                                log_fn=lambda m: self._log(f"{tag}{m}", "muted")),
                            timeout=300)
                        _addqt_ok = bool(oka)
                        if oka:
                            self._log(f"{tag} ✅ {msga}", "ok")
                            task_detail = (task_detail + f" | Add QT TH: {msga}").strip(" |")
                        elif _codea == "2FA7D":
                            # Google chặn xác minh (We couldn't verify it's you) → note '2FA 7 ngày'
                            _addqt_note = "2FA 7 ngày"
                            self._log(f"{tag} ⛔ Add QT TH: 2FA 7 ngày (Google chặn xác minh)", "warn")
                            task_detail = (task_detail + " | Add QT TH: 2FA 7 ngày").strip(" |")
                        else:
                            self._log(f"{tag} ⚠ Add QT TH: {msga}", "warn")
                            task_detail = (task_detail + f" | Add QT TH lỗi: {msga}").strip(" |")
                    except asyncio.TimeoutError:
                        _addqt_ok = False
                        self._log(f"{tag} ✗ Add QT TH QUÁ 5 PHÚT — bỏ qua.", "err")
                        task_detail = (task_detail + " | Add QT TH: quá giờ").strip(" |")
                    except Exception as ea:
                        _addqt_ok = False
                        self._log(f"{tag} ✗ Lỗi Add QT TH: {ea}", "err")
                        task_detail = (task_detail + f" | Lỗi Add QT TH: {ea}").strip(" |")

                # ── ADD THÊM QT: CHỈ add thêm 1 mail owner vào brand account có sẵn ──
                # (bỏ qua tạo kênh + chuyển kênh — skip_create_move=True)
                async def _run_addqtonly():
                    nonlocal _addqtonly_ok, _addqtonly_note, task_detail
                    self._log(f"{tag} ➕ Add Thêm QT…", "muted")
                    self._tree_upsert(email, status="⏳ Add Thêm QT…", tag="run")
                    try:
                        oka, _codea, msga = await asyncio.wait_for(
                            do_add_brand_admin(
                                ws_url, email, password=acc.get("password", ""),
                                owner_email=acc.get("owner", ""),   # cột E: mail quản trị
                                totp_secret=(new_totp or acc.get("totp", "")),
                                skip_create_move=True,              # CHỈ add owner
                                log_fn=lambda m: self._log(f"{tag}{m}", "muted")),
                            timeout=420)
                        _addqtonly_ok = bool(oka)
                        if oka:
                            self._log(f"{tag} ✅ {msga}", "ok")
                            task_detail = (task_detail + f" | Add Thêm QT: {msga}").strip(" |")
                        elif _codea == "2FA7D":
                            _addqtonly_note = "2FA 7 ngày"
                            self._log(f"{tag} ⛔ Add Thêm QT: 2FA 7 ngày (Google chặn xác minh)", "warn")
                            task_detail = (task_detail + " | Add Thêm QT: 2FA 7 ngày").strip(" |")
                        else:
                            self._log(f"{tag} ⚠ Add Thêm QT: {msga}", "warn")
                            task_detail = (task_detail + f" | Add Thêm QT lỗi: {msga}").strip(" |")
                    except asyncio.TimeoutError:
                        _addqtonly_ok = False
                        self._log(f"{tag} ✗ Add Thêm QT QUÁ 5 PHÚT — bỏ qua.", "err")
                        task_detail = (task_detail + " | Add Thêm QT: quá giờ").strip(" |")
                    except Exception as eao:
                        _addqtonly_ok = False
                        self._log(f"{tag} ✗ Lỗi Add Thêm QT: {eao}", "err")
                        task_detail = (task_detail + f" | Lỗi Add Thêm QT: {eao}").strip(" |")

                # ── CHẤP NHẬN QT: vào Brand Accounts → chấp nhận TẤT CẢ lời mời quản trị ──
                async def _run_cnqt():
                    nonlocal _cnqt_ok, _cnqt_note, task_detail
                    self._log(f"{tag} ✅ Chấp nhận QT…", "muted")
                    self._tree_upsert(email, status="⏳ Chấp nhận QT…", tag="run")
                    try:
                        okc2, _codec2, msgc2 = await asyncio.wait_for(
                            do_accept_brand_invite(
                                ws_url, email, password=acc.get("password", ""),
                                totp_secret=(new_totp or acc.get("totp", "")),
                                log_fn=lambda m: self._log(f"{tag}{m}", "muted")),
                            timeout=300)
                        _cnqt_ok = bool(okc2)
                        if okc2:
                            self._log(f"{tag} ✅ {msgc2}", "ok")
                            task_detail = (task_detail + f" | Chấp nhận QT: {msgc2}").strip(" |")
                        elif _codec2 == "2FA7D":
                            _cnqt_note = "2FA 7 ngày"
                            self._log(f"{tag} ⛔ Chấp nhận QT: 2FA 7 ngày", "warn")
                            task_detail = (task_detail + " | Chấp nhận QT: 2FA 7 ngày").strip(" |")
                        else:
                            self._log(f"{tag} ⚠ Chấp nhận QT: {msgc2}", "warn")
                            task_detail = (task_detail + f" | Chấp nhận QT lỗi: {msgc2}").strip(" |")
                    except asyncio.TimeoutError:
                        _cnqt_ok = False
                        self._log(f"{tag} ✗ Chấp nhận QT QUÁ 5 PHÚT — bỏ qua.", "err")
                        task_detail = (task_detail + " | Chấp nhận QT: quá giờ").strip(" |")
                    except Exception as ec2:
                        _cnqt_ok = False
                        self._log(f"{tag} ✗ Lỗi Chấp nhận QT: {ec2}", "err")
                        task_detail = (task_detail + f" | Lỗi Chấp nhận QT: {ec2}").strip(" |")

                # ── TẠO DATA: đổi bìa + đổi avatar + chụp thông tin kênh ──
                # Lỗi 1 bước VẪN chạy tiếp bước còn lại, ghi rõ bước nào lỗi.
                async def _run_taodata():
                    nonlocal _taodata_ok, _taodata_detail, task_detail
                    self._log(f"{tag} 🗂️ Tạo Data…", "muted")
                    self._tree_upsert(email, status="⏳ Tạo Data…", tag="run")
                    # Mỗi kênh 1 folder riêng: DATA_KENH/<email>/
                    _data_root = Path(__file__).parent.parent / "DATA_KENH"
                    _acc_dir = _data_root / (email.replace("@", "_at_").replace("/", "_") or "unknown")
                    try:
                        _acc_dir.mkdir(parents=True, exist_ok=True)
                    except Exception:
                        pass
                    _bdir2 = str(Path(__file__).parent / "anh_bia")

                    # BÌA + AVATAR trong CÙNG 1 phiên Studio (avatar search tên KÊNH THỨ 2).
                    # Lỗi 1 bước vẫn làm bước còn lại; detail ghi rõ bước nào OK/lỗi.
                    try:
                        okd, _, msgd = await asyncio.wait_for(
                            do_taodata_images(
                                ws_url, email, banner_dir=_bdir2, data_dir=str(_acc_dir),
                                log_fn=lambda m: self._log(f"{tag}{m}", "muted")),
                            timeout=480)
                        _taodata_ok = bool(okd)
                        _taodata_detail = msgd
                        if okd:
                            self._log(f"{tag} ✅ (Tạo Data) {msgd}", "ok")
                        else:
                            self._log(f"{tag} ⚠ (Tạo Data) {msgd}", "warn")
                    except asyncio.TimeoutError:
                        _taodata_ok = False
                        _taodata_detail = "quá giờ (8 phút)"
                        self._log(f"{tag} ✗ (Tạo Data) quá 8 phút — bỏ qua.", "err")
                    except Exception as ed:
                        _taodata_ok = False
                        _taodata_detail = str(ed)
                        self._log(f"{tag} ✗ (Tạo Data) lỗi: {ed}", "err")

                    task_detail = (task_detail + f" | Tạo Data: {_taodata_detail}").strip(" |")

                # ══ GỌI các tác vụ theo ĐÚNG THỨ TỰ NGƯỜI DÙNG CLICK ══
                # (create2fa/change2fa gộp = '2fa'; channel/getlink gộp = 'channel' → chạy 1 lần)
                _dispatch = {
                    "create2fa": ("2fa", _run_2fa),          "change2fa": ("2fa", _run_2fa),
                    "channel":   ("channel", _run_channel),  "getlink":   ("channel", _run_channel),
                    "banner":    ("banner", _run_banner),
                    "rmadmin":   ("rmadmin", _run_rmadmin),
                    "addqtth":   ("addqtth", _run_addqtth),
                    "addqtonly": ("addqtonly", _run_addqtonly),
                    "cnqt":      ("cnqt", _run_cnqt),
                    "taodata":   ("taodata", _run_taodata),
                }
                _run_order = [t for t in _sel if t != "login" and t in _dispatch]
                if _run_order:
                    self._log(f"{tag} 🧭 Thứ tự chạy: {' → '.join(_run_order)}", "muted")
                _ran_keys = set()
                for _tk in _run_order:
                    if self.stop_flag:
                        break
                    _key, _fn = _dispatch[_tk]
                    if _key in _ran_keys:
                        continue
                    _ran_keys.add(_key)
                    await _fn()

            # ⚠ LƯU 2FA MỚI TRƯỚC MỌI NHÁNH GHI KẾT QUẢ — nếu Tạo/Đổi 2FA đã sinh secret mới
            # thì PHẢI ghi nó vào creds dù nhiệm vụ con SAU đó (bìa/thoát QT/Add QT TH) thất bại.
            # (Lỗi cũ: các nhánh THẤT BẠI gọi _record mà chưa lưu new_totp → MẤT 2FA mới.)
            if new_totp:
                self._acc_creds.setdefault(email, acc)["totp"] = new_totp

            # Đăng nhập OK nhưng 1 NHIỆM VỤ CON THẤT BẠI → KHÔNG được báo "THÀNH CÔNG".
            # (Trước đây 'success' chỉ phản ánh bước login nên nhiệm vụ con lỗi vẫn ghi OK.)
            if self.stop_flag:
                # Bấm DỪNG giữa chừng → account có thể CHƯA chạy hết tác vụ. Ghi 'ĐÃ DỪNG'
                # và KHÔNG đánh dấu hoàn tất (xem _test_after_record) để resume CHẠY LẠI account này.
                self._log(f"{tag} ⏸ Đã DỪNG — {email} có thể chưa xong hết (sẽ chạy lại khi resume).", "warn")
                self._record(key, email, "ĐÃ DỪNG (chưa xong hết)", ip_str,
                             task_detail, channel or "")
            elif success and _banner_ok is False:
                self._log(f"{tag} ⚠ THAY ẢNH BÌA THẤT BẠI: {email}", "warn")
                self._record(key, email, "THAY ẢNH BÌA THẤT BẠI", ip_str,
                             task_detail, channel or "")
            elif success and _rmadmin_ok is False:
                self._log(f"{tag} ⚠ THOÁT QUẢN TRỊ THẤT BẠI: {email}", "warn")
                self._record(key, email, "THOÁT QUẢN TRỊ THẤT BẠI", ip_str,
                             task_detail, channel or "")
            elif success and _addqt_ok is False and _addqt_note == "2FA 7 ngày":
                self._log(f"{tag} ⛔ 2FA 7 NGÀY: {email}", "warn")
                self._record(key, email, "2FA 7 ngày", ip_str,
                             task_detail, channel or "")
            elif success and _addqt_ok is False:
                self._log(f"{tag} ⚠ ADD QT TH THẤT BẠI: {email}", "warn")
                self._record(key, email, "ADD QT TH THẤT BẠI", ip_str,
                             task_detail, channel or "")
            elif success and _addqtonly_ok is False and _addqtonly_note == "2FA 7 ngày":
                self._log(f"{tag} ⛔ 2FA 7 NGÀY (Add Thêm QT): {email}", "warn")
                self._record(key, email, "2FA 7 ngày", ip_str, task_detail, channel or "")
            elif success and _addqtonly_ok is False:
                self._log(f"{tag} ⚠ ADD THÊM QT THẤT BẠI: {email}", "warn")
                self._record(key, email, "ADD THÊM QT THẤT BẠI", ip_str,
                             task_detail, channel or "")
            elif success and _cnqt_ok is False and _cnqt_note == "2FA 7 ngày":
                self._log(f"{tag} ⛔ 2FA 7 NGÀY (Chấp nhận QT): {email}", "warn")
                self._record(key, email, "2FA 7 ngày", ip_str, task_detail, channel or "")
            elif success and _cnqt_ok is False:
                self._log(f"{tag} ⚠ CHẤP NHẬN QT THẤT BẠI: {email}", "warn")
                self._record(key, email, "CHẤP NHẬN QT THẤT BẠI", ip_str,
                             task_detail, channel or "")
            elif success and _taodata_ok is False:
                self._log(f"{tag} ⚠ TẠO DATA LỖI 1 PHẦN: {email} ({_taodata_detail})", "warn")
                self._record(key, email, f"TẠO DATA LỖI ({_taodata_detail})", ip_str,
                             task_detail, channel or "")
            elif success:
                extra = f" | Kênh: {channel}" if channel else ""
                self._log(f"{tag} ✅ {task_status}: {email}{extra}", "ok")
                self._record(key, email, task_status, ip_str, task_detail, channel or "")
            elif self._is_proxy_error(err_str):
                self._log(f"{tag} ✗ Proxy chết (đã đổi IP nhiều lần): {err_str[:60]}", "err")
                self._record(key, email, "LỖI PROXY (mạng)", ip_str, err_str[:150])
            else:
                st, dt = self._classify(err_str)
                self._log(f"{tag} ⚠ {st}: {err_str[:80]}", "warn")
                self._record(key, email, st, ip_str, dt)

            # 7) Đóng browser nếu không giữ mở -----------------------------
            if not self.keep_open_var.get():
                gpm_stop(profile_id)
                self._log(f"{tag} Browser đã đóng", "muted")

            # 8) Nghỉ 20s trước khi tạo profile + chạy kênh kế (trừ kênh cuối)
            if not self._acc_queue.empty() and not self.stop_flag:
                self._log(f"{tag} ⏸ Xong kênh — nghỉ 20s trước khi sang kênh kế…", "muted")
                self._set_status(f"{tag} nghỉ 20s trước kênh kế…")
                self._sleep_stop(20, tag)

        self._log(f"{tag} ✔ Xong luồng.", "ok")

    # ── Kết quả login ──────────────────────────────────────
    def _record(self, key: str, email: str, status: str,
                ip: str = "", detail: str = "", channel: str = ""):
        creds = getattr(self, "_acc_creds", {}).get(email, {})
        _pw   = creds.get("password", "")
        _rec  = creds.get("recovery", "")
        _totp = creds.get("totp", "")   # đã là 2FA MỚI nếu vừa tạo/đổi
        row = {
            "time": time.strftime("%Y-%m-%d %H:%M:%S"),
            "key": key,
            "email": email,
            "password": _pw,
            "recovery": _rec,
            "totp": _totp,
            "old_totp": creds.get("old_totp", ""),
            "ip": ip,
            "status": status,
            "detail": (detail or "")[:200],
            "channel": channel or "",
            # Dòng login gộp — KHÔNG khoảng trắng quanh dấu '|', dùng 2FA mới
            "login": f"{email}|{_pw}|{_rec}|{_totp}",
        }
        with self._results_lock:
            self.results.append(row)
        # Ghi kết quả lên Google Sheet (nếu bật) — chạy trong worker thread
        try:
            self._gsheet_push(row)
        except Exception:
            pass
        # Cập nhật progress (khóa: += không atomic khi đa luồng)
        with self._progress_lock:
            self._progress_done += 1
            _done, _total = self._progress_done, self._progress_total
        self.root.after(0, lambda d=_done, t=_total: [
            self.progress_lbl.config(text=f"{d} / {t} tài khoản"),
            self.progress_lbl2.config(text=f"{d} / {t} tài khoản")])
        # Cập nhật bảng kết quả
        su = status.upper()
        tag = "ok" if "THÀNH CÔNG" in su else (
            "warn" if ("XÁC MINH" in su or "CHƯA RÕ" in su) else "err")
        self._tree_upsert(email, ip=ip or None, status=status, detail=detail, tag=tag)
        # CHẾ ĐỘ TEST: đánh dấu done/skip + dừng khi lỗi
        try:
            self._test_after_record(email, status, detail, channel or "")
        except Exception:
            pass

    @staticmethod
    def _is_proxy_error(msg: str) -> bool:
        """True nếu lỗi do proxy không có mạng / kết nối proxy thất bại."""
        m = (msg or "").lower()
        keys = [
            "err_proxy_connection_failed", "err_tunnel_connection_failed",
            "err_socks_connection_failed", "err_no_supported_proxies",
            "err_connection_timed_out", "err_timed_out", "err_connection_reset",
            "err_connection_closed", "err_connection_refused",
            "err_empty_response", "err_name_not_resolved", "err_address_unreachable",
            "err_internet_disconnected", "net::err", "proxy_connection",
            "proxy lỗi", "timeout", "timederror",
            # Lỗi điều hướng trang do proxy chậm/rớt → nên đổi IP chạy lại
            "page.goto", "navigation to", "navigation failed", "err_aborted",
            "ns_binding_aborted", "frame was detached",
        ]
        return any(k in m for k in keys)

    @staticmethod
    def _classify(result_msg: str) -> tuple[str, str]:
        """Phân loại kết quả login thất bại -> (trạng thái, chi tiết)."""
        m = (result_msg or "").lower()
        # CAPTCHA
        if any(k in m for k in ["captcha |", "captcha", "không phải là người máy",
                                 "not a robot", "verify your identity", "xác minh danh tính"]):
            return "GOOGLE CHẶN: CAPTCHA (đổi proxy)", result_msg
        # Mật khẩu sai hoặc đã thay đổi (/challenge/pwd)
        # Phải check TRƯỚC "quá nhiều lần" vì message có thể chứa cả hai
        if any(k in m for k in ["sai_mk |", "challenge/pwd", "mật khẩu sai hoặc đã thay đổi",
                                 "xác nhận lại mật khẩu"]):
            return "MẬT KHẨU SAI HOẶC ĐÃ ĐỔI (/challenge/pwd)", result_msg
        # Sai mã 2FA — check TRƯỚC "quá nhiều lần" vì SAI_2FA hay kèm "quá nhiều lần thử"
        if any(k in m for k in ["sai_2fa", "sal_2fa", "wrong code", "sai mã", "mã không đúng",
                                 "không đúng mã", "invalid code", "incorrect code",
                                 "code you entered", "mã xác minh không",
                                 "mã 2fa không đúng", "mã 2fa"]):
            return "SAI MÃ 2FA", result_msg
        # Google chặn: quá nhiều lần thử
        if any(k in m for k in ["too_many_attempts", "too many failed", "too many attempts",
                                 "quá nhiều lần", "try again later", "thử lại sau",
                                 "unusual traffic"]):
            return "GOOGLE CHẶN: QUÁ NHIỀU LẦN THỬ (thử lại sau)", result_msg
        # Google chặn: trình duyệt không an toàn
        if any(k in m for k in ["browser_not_secure", "not secure", "browser or app",
                                 "không an toàn"]):
            return "GOOGLE CHẶN: TRÌNH DUYỆT KHÔNG AN TOÀN", result_msg
        # Tài khoản bị vô hiệu hoá / khoá
        if any(k in m for k in ["disabled", "vô hiệu", "bị tắt", "bị khoá", "bị khóa",
                                 "account has been disabled", "tài khoản đã bị",
                                 "suspended", "terminated", "bị vô hiệu hoá"]):
            return "GMAIL BỊ VÔ HIỆU HOÁ (disabled)", result_msg
        # Sai mật khẩu (Google thông báo rõ wrong password)
        if any(k in m for k in ["sai_matkhau", "wrong password", "sai mật khẩu",
                                 "mật khẩu bạn đã nhập không đúng",
                                 "incorrect password", "password you entered"]):
            return "SAI MẬT KHẨU", result_msg
        # Sai email khôi phục (recovery)
        if any(k in m for k in ["sai_recovery", "recovery email", "email khôi phục",
                                 "khôi phục không", "recovery không đúng",
                                 "confirm the recovery"]):
            return "SAI EMAIL KHÔI PHỤC", result_msg
        # Sai email / không tồn tại
        if any(k in m for k in ["sai_email", "couldn't find", "không tìm thấy tài khoản",
                                 "ô nhập email", "no account found",
                                 "không tìm thấy tài khoản google"]):
            return "SAI EMAIL / KHÔNG TỒN TẠI", result_msg
        # Cần KIỂM TRA ĐIỆN THOẠI (Google gửi thông báo đẩy lên máy, bấm số để duyệt — tool
        # không tự làm được). Đặt TRƯỚC 'verify' vì message chứa 'phone' dễ bị nuốt.
        if any(k in m for k in ["phone_check", "kiểm tra điện thoại", "check your phone",
                                 "thông báo đến điện thoại", "trên điện thoại để xác minh"]):
            return "KIỂM TRA ĐIỆN THOẠI (duyệt tay trên máy)", result_msg
        # Lỗi TẢI/ĐIỀU HƯỚNG trang (page.goto timeout / navigation huỷ / net::err) — do proxy
        # chậm/rớt mạng, KHÔNG phải account lỗi. Đặt TRƯỚC 'verify' vì URL trong message có thể
        # chứa 'challenge/signin' → dễ phân loại nhầm thành 'cần xác minh'.
        if any(k in m for k in ["page.goto", "navigation to", "navigation failed",
                                 "err_aborted", "ns_binding_aborted", "net::err",
                                 "frame was detached", "load event", "timeout 3"]):
            return "LỖI TẢI TRANG (proxy/mạng — chạy lại)", result_msg
        # Cần xác minh (verify: điện thoại / thiết bị lạ)
        if any(k in m for k in ["verify", "verification", "xác minh", "challenge",
                                 "unusual activity", "hoạt động bất thường",
                                 "confirm it", "xác nhận danh tính", "phone",
                                 "số điện thoại", "2fa", "2-step"]):
            return "CẦN XÁC MINH (verify/2FA/điện thoại)", result_msg
        if "không tìm thấy ô nhập mật khẩu" in m:
            return "KHÔNG VÀO ĐƯỢC BƯỚC MẬT KHẨU", result_msg
        return "CHƯA RÕ — KIỂM TRA TAY", result_msg

    def _save_results(self, open_after: bool = False):
        if not self.results:
            self._log("• Chưa có kết quả nào để xuất.", "muted")
            return None
        # Bật Google Sheet + 'không tải Excel local' → bỏ qua xuất file (chỉ khi tự lưu).
        if (not open_after and self._gsheet_on()
                and bool(getattr(self, "gsheet_nolocal_var", None) and self.gsheet_nolocal_var.get())):
            self._log("• Đã ghi kết quả lên Google Sheet (không tải Excel local).", "muted")
            return None
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            self._log("⚠ Chưa cài openpyxl, đang cài...", "warn")
            import subprocess, sys
            subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl",
                            "--break-system-packages", "-q"], check=False)
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

        folder = Path(__file__).parent / "data"
        folder.mkdir(exist_ok=True)
        path = folder / f"ket_qua_login_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
        cols    = ["time", "key", "email", "password", "recovery", "totp", "old_totp",
                   "ip", "status", "detail", "channel", "login"]
        headers = ["Th\u1eddi gian", "Key", "Email", "M\u1eadt kh\u1ea9u", "Recovery mail",
                   "2FA secret", "2FA c\u0169", "IP proxy", "Tr\u1ea1ng th\u00e1i", "Chi ti\u1ebft",
                   "Link k\u00eanh", "Login (mail|pass|rec|2fa)"]
        col_widths = [18, 12, 32, 18, 28, 20, 20, 30, 32, 55, 28, 60]

        STATUS_COLOR = {
            "TH\u00c0NH C\u00d4NG":               "00B050",
            "SAI M\u00c3 2FA":               "FF6600",
            "GOOGLE CH\u1eb6N: CAPTCHA":     "C00000",
            "GOOGLE CH\u1eb6N: QU\u00c1 NHI\u1ec0U":   "C00000",
            "CH\u01af\u0410 R\u00d5":                  "FFC000",
            "C\u1ea6N X\u00c1C MINH":             "0070C0",
        }
        def _status_color(st):
            for k, v in STATUS_COLOR.items():
                if k in st:
                    return v
            return "FF0000"

        try:
            with self._results_lock:
                rows = list(self.results)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "K\u1ebft qu\u1ea3"

            hdr_font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
            hdr_fill  = PatternFill("solid", fgColor="2F5597")
            hdr_align = Alignment(horizontal="center", vertical="center")

            for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
                c = ws.cell(1, ci, h)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = hdr_align
                ws.column_dimensions[get_column_letter(ci)].width = w
            ws.row_dimensions[1].height = 22

            for ri, r in enumerate(rows, 2):
                st  = r.get("status", "")
                clr = _status_color(st)
                for ci, col in enumerate(cols, 1):
                    val = r.get(col, "") or ""
                    cell = ws.cell(ri, ci, val)
                    if ci == 9:  # Tr\u1ea1ng th\u00e1i (\u0111\u00e3 th\u00eam c\u1ed9t 2FA c\u0169 n\u00ean d\u1ecbch sang 9)
                        cell.font = Font(name="Arial", size=10, color=clr, bold=True)
                    else:
                        cell.font = Font(name="Arial", size=10, color="000000")
                    cell.alignment = Alignment(vertical="center", wrap_text=False)

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

            wb.save(path)
            self._log(f"\u2714 L\u01b0u k\u1ebft qu\u1ea3: {path.name} ({len(rows)} d\u00f2ng)", "ok")
            if open_after:
                try:
                    os.startfile(str(path))
                except Exception:
                    pass
            return path
        except Exception as e:
            self._log(f"\u2717 L\u01b0u k\u1ebft qu\u1ea3: {e}", "err")
            self._log(traceback.format_exc()[:300], "err")
            return None


if __name__ == "__main__":
    app = KiotLoginApp()
    app.root.mainloop()
